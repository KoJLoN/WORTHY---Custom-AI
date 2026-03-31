"""Worthy.py — unified WORTHY SYSTEM mega-engine.

This file embeds six original engines:

W — Writer
O — Operations / Control-Sheet
R — Research / Mission Planner
T — Technical Engineer / Iterative Solution Engineering
H — Life Planner / Psychometric System
Y — Persona / YOU-Mode Engine

Each engine's original source code is stored as a raw string constant and
executed in an isolated namespace when that engine is selected.
"""

from pathlib import Path
import openpyxl

# ------------------------------------------------------------
# QUIET MODE — Silence HTTP request spam + OpenAI logs
# ------------------------------------------------------------
import logging
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("openai").setLevel(logging.WARNING)


WORTHY_XLSX_PATH = "Worthy.xlsx"


def get_selector_and_background():
    """
    Read Sheet1!A1 as selector (W/O/R/T/H/Y)
    and Sheet1!A3 downward as background text.
    """
    path = Path(WORTHY_XLSX_PATH)
    if not path.exists():
        raise SystemExit(f"Worthy workbook not found: {path!s}")

    wb = openpyxl.load_workbook(path)
    if "Sheet1" not in wb.sheetnames:
        raise SystemExit("Sheet1 not found in Worthy.xlsx")

    ws = wb["Sheet1"]
    selector_val = ws["A1"].value or ""
    selector = str(selector_val).strip().upper()

    # A3 downward = background text
    lines = []
    row = 3
    while True:
        val = ws.cell(row=row, column=1).value
        if val is None:
            break
        lines.append(str(val))
        row += 1

    background = "\n".join(lines).strip()
    return selector, background


# ---------------------------------------------------------------------------
# Embedded engine source code (as raw strings)
# ---------------------------------------------------------------------------

# ===== Engine W source =====
#===ENGINE_W_START===
W_CODE = r'''#!/usr/bin/env python3
"""
local_writer.py

Fully autonomous local writing engine:
- Reads Sheet1 description
- Generates outline (Sheet2)
- Builds prompts (Sheet3)
- Calls GPT-4.1-mini via Responses API
- Writes full prose (NO bullet points)
- Saves after every section
- Bash-friendly progress output
- Dynamic temperature
- Uses Worthy.xlsx as the primary workbook
"""

import sys
import random
import time
import re
from pathlib import Path
from typing import List, Dict

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openai import OpenAI

client = OpenAI()

MODEL_NAME = "gpt-4.1-mini"

SHEET1 = "Sheet1"
SHEET2 = "Sheet2"
SHEET3 = "Sheet3"
SHEET5 = "Sheet5"

BULLET_PATTERN = re.compile(r"^\s*[-•*]", re.MULTILINE)

SUPPORTED_TYPES = [
    "Fiction Novel", "Short story", "script", "poem", "personal essay", "biography",
    "creative writing", "news article", "research paper", "dissertation", "grant proposal",
    "technical report", "conference paper", "textbook", "how to book", "encylopedia",
    "instruction manual", "business letter", "cover letter", "pitch deck", "nonfiction novel",
    "social media post", "speech"
]

STYLE = {
    "Fiction Novel": {"pages": (200, 350), "sections": (15, 40), "pp_page": 300,
                      "subrange": (8, 20), "struct": "chapters"},
    "nonfiction novel": {"pages": (200, 350), "sections": (15, 35), "pp_page": 280,
                         "subrange": (8, 18), "struct": "chapters"},
    "Short story": {"pages": (5, 40), "sections": (3, 8), "pp_page": 275, "struct": "sections"},
    "script": {"pages": (30, 120), "sections": (10, 60), "pp_page": 180, "struct": "scenes"},
    "poem": {"pages": (1, 5), "sections": (3, 20), "pp_page": 150, "struct": "stanzas"},
    "personal essay": {"pages": (3, 15), "sections": (3, 6), "pp_page": 275, "struct": "sections"},
    "biography": {"pages": (150, 400), "sections": (15, 40), "pp_page": 300,
                  "subrange": (8, 25), "struct": "chapters"},
    "creative writing": {"pages": (5, 100), "sections": (4, 20), "pp_page": 280, "struct": "chapters"},
    "news article": {"pages": (1, 5), "sections": (3, 7), "pp_page": 250, "struct": "sections"},
    "research paper": {"pages": (8, 40), "sections": (5, 12), "pp_page": 275, "struct": "sections"},
    "dissertation": {"pages": (150, 400), "sections": (5, 12), "pp_page": 275,
                     "subrange": (15, 40), "struct": "chapters"},
    "grant proposal": {"pages": (5, 30), "sections": (5, 10), "pp_page": 260, "struct": "sections"},
    "technical report": {"pages": (10, 80), "sections": (5, 15), "pp_page": 280, "struct": "sections"},
    "conference paper": {"pages": (6, 15), "sections": (5, 9), "pp_page": 275, "struct": "sections"},
    "textbook": {"pages": (200, 800), "sections": (8, 25), "pp_page": 300,
                 "subrange": (10, 40), "struct": "chapters"},
    "how to book": {"pages": (120, 300), "sections": (10, 30), "pp_page": 275,
                    "subrange": (8, 20), "struct": "chapters"},
    "encylopedia": {"pages": (50, 600), "sections": (20, 200), "pp_page": 260, "struct": "entries"},
    "instruction manual": {"pages": (10, 200), "sections": (5, 40), "pp_page": 260, "struct": "sections"},
    "business letter": {"pages": (1, 2), "sections": (1, 3), "pp_page": 250, "struct": "sections"},
    "cover letter": {"pages": (1, 2), "sections": (1, 3), "pp_page": 250, "struct": "sections"},
    "pitch deck": {"pages": (10, 30), "sections": (10, 30), "pp_page": 100, "struct": "slides"},
    "social media post": {"pages": (.25, 1), "sections": (1, 3), "pp_page": 200, "struct": "sections"},
    "speech": {"pages": (2, 20), "sections": (4, 12), "pp_page": 275, "struct": "sections"},
}

### ------------------------------------------------------------
### Utilities
### ------------------------------------------------------------

def get_sheet(wb, name):
    return wb[name] if name in wb.sheetnames else wb.create_sheet(name)

def read_contiguous(sheet, col="A"):
    vals = []
    row = 1
    while True:
        v = sheet[f"{col}{row}"].value
        if v is None or str(v).strip() == "":
            break
        vals.append(str(v))
        row += 1
    return "\n".join(vals).strip()

def quant25(x): return round(x * 4) / 4.0

### ------------------------------------------------------------
### Dynamic temperature
### ------------------------------------------------------------

def infer_temperature(desc: str) -> float:
    d = desc.lower()
    if any(k in d for k in ["research", "academic", "dissertation", "policy", "technical", "methodology"]):
        return 0.2
    if any(k in d for k in ["novel", "story", "creative", "fiction", "poetic", "imaginative"]):
        return 0.7
    return 0.4

### ------------------------------------------------------------
### Writing type inference
### ------------------------------------------------------------

def infer_type(sheet2, desc):
    if sheet2["A1"].value in SUPPORTED_TYPES:
        return sheet2["A1"].value
    d = desc.lower()
    for t in SUPPORTED_TYPES:
        if t.lower().split()[0] in d:
            return t
    if "novel" in d: return "Fiction Novel"
    return "creative writing"

### ------------------------------------------------------------
### Outline generation
### ------------------------------------------------------------

def global_plan(wtype):
    c = STYLE[wtype]
    pages = quant25(random.uniform(*c["pages"]))
    pp_page = c["pp_page"]
    wpp = 250 if pages <= 5 else 350
    secs = random.randint(*c["sections"])
    return {
        "type": wtype,
        "pages": pages,
        "pp_page": pp_page,
        "wpp": wpp,
        "sections": secs,
        "struct": c["struct"],
        "sub": c.get("subrange", None)
    }

def distribute_pages(plan):
    n = plan["sections"]
    if plan["sub"]:
        raw = [random.uniform(*plan["sub"]) for _ in range(n)]
    else:
        avg = plan["pages"] / n
        raw = [random.uniform(0.6 * avg, 1.4 * avg) for _ in range(n)]
    scale = plan["pages"] / sum(raw)
    return [quant25(r * scale) for r in raw]

def build_outline(sheet2, plan):
    sheet2["A1"] = plan["type"]
    sheet2["A2"] = "SectionIndex"
    sheet2["B2"] = "Label"
    sheet2["C2"] = "Structure"
    sheet2["D2"] = "Pages"
    sheet2["E2"] = "Words"
    sheet2["F2"] = "Prompts"
    sheet2["G2"] = "Completed"

    pages = distribute_pages(plan)
    row = 3

    print("[INIT] Building outline...")

    for i, p in enumerate(pages, 1):
        words = int(p * plan["pp_page"])
        prompts = max(1, math.ceil(words / plan["wpp"]))
        label = f"{plan['struct'][:-1].capitalize()} {i}"

        sheet2[f"A{row}"] = i
        sheet2[f"B{row}"] = label
        sheet2[f"C{row}"] = plan["struct"]
        sheet2[f"D{row}"] = p
        sheet2[f"E{row}"] = words
        sheet2[f"F{row}"] = prompts
        sheet2[f"G{row}"] = 0
        row += 1

### ------------------------------------------------------------
### Load outline
### ------------------------------------------------------------

def load_outline(sheet2):
    out = []
    for r in range(3, sheet2.max_row + 1):
        idx = sheet2[f"A{r}"].value
        if idx is None:
            continue
        out.append({
            "index": int(idx),
            "label": sheet2[f"B{r}"].value,
            "struct": sheet2[f"C{r}"].value,
            "pages": float(sheet2[f"D{r}"].value),
            "words": int(sheet2[f"E{r}"].value),
            "prompts": int(sheet2[f"F{r}"].value),
            "completed": int(sheet2[f"G{r}"].value),
            "row": r
        })
    return out

### ------------------------------------------------------------
### Prompt generator
### ------------------------------------------------------------

def make_prompt(desc, wtype, sec, pidx, total):
    approx = sec["words"] // total

    return f"""
Continue the {wtype.lower()} in continuous narrative or expository prose—NO bullet points, NO lists.

Project description:
\"\"\"{desc}\"\"\"

You are working on {sec['struct'][:-1]} "{sec['label']}".

This is prompt {pidx} of {total}.
Continue directly from where the previous prompt ended.

Write approximately {approx} words of smooth, uninterrupted prose.
"""

### ------------------------------------------------------------
### GPT Call
### ------------------------------------------------------------

def call_llm(prompt, temp):
    print("[PROMPT] Calling GPT-4.1-mini...")
    resp = client.responses.create(
        model=MODEL_NAME,
        input=prompt,
        temperature=temp
    )
    text = resp.output[0].content[0].text.strip()
    if BULLET_PATTERN.search(text):
        text = re.sub(BULLET_PATTERN, "", text)
    return text.strip()

### ------------------------------------------------------------
### Write one full section
### ------------------------------------------------------------

def write_section(wb, sheet3, sheet2, sec, desc, wtype, temp, total_prompts_global):
    section_total = sec["prompts"]
    idx = sec["index"]
    label = sec["label"]

    print(f"\n[SECTION] Starting {label} (Section {idx})")
    print(f"[SECTION] Pages: {sec['pages']} | Words: {sec['words']} | Prompts: {section_total}")

    row = sheet3.max_row + 1

    for pidx in range(sec["completed"] + 1, section_total + 1):

        written_count = sum(s["completed"] for s in total_prompts_global)
        remaining = sum(s["prompts"] for s in total_prompts_global) - written_count

        print(f"[PROMPT] {label} — Prompt {pidx}/{section_total}")
        print(f"[PROGRESS] Total prompts completed: {written_count} | Remaining: {remaining}")

        pr = make_prompt(desc, wtype, sec, pidx, section_total)
        res = call_llm(pr, temp)

        sheet3[f"A{row}"] = pr
        sheet3[f"B{row}"] = res
        sheet3[f"C{row}"] = idx
        sheet3[f"D{row}"] = pidx
        sheet3[f"E{row}"] = sec["words"] // section_total
        row += 1

        # Update section progress
        sec["completed"] = pidx
        sheet2[f"G{sec['row']}"] = pidx

        wb.save(workbook_path)
        print(f"[PROMPT COMPLETE] Finished prompt {pidx}/{section_total}")

    print(f"[SECTION COMPLETE] {label} fully written and saved.")

### ------------------------------------------------------------
### MAIN
### ------------------------------------------------------------

if __name__ == "__main__":

    if len(sys.argv) < 2:
        print("Usage: python local_writer.py <Worthy.xlsx>")
        print("Default: python local_writer.py Worthy.xlsx")
        sys.exit(1)

    workbook_path = Path(sys.argv[1]).resolve()

    # If user didn't provide a file, assume Worthy.xlsx
    if not workbook_path.exists() and sys.argv[1] == "Worthy.xlsx":
        print("[INIT] Creating new workbook: Worthy.xlsx")
        wb = openpyxl.Workbook()
        wb.active.title = SHEET1
        wb.create_sheet(SHEET2)
        wb.create_sheet(SHEET3)
        wb.create_sheet(SHEET5)
        wb.save(workbook_path)

    elif not workbook_path.exists():
        print(f"ERROR: Workbook not found: {workbook_path}")
        sys.exit(1)

    print(f"[INIT] Using workbook: {workbook_path.name}")

    wb = openpyxl.load_workbook(workbook_path)
    sheet1 = get_sheet(wb, SHEET1)
    sheet2 = get_sheet(wb, SHEET2)
    sheet3 = get_sheet(wb, SHEET3)
    sheet5 = get_sheet(wb, SHEET5)

    print("[INIT] Loading project from Worthy.xlsx...")

    desc = read_contiguous(sheet1)
    if not desc:
        print("ERROR: Sheet1 is empty.")
        sys.exit(1)

    temp = infer_temperature(desc)
    print(f"[INIT] Inferred temperature: {temp}")

    wtype = infer_type(sheet2, desc)
    print(f"[INIT] Writing type detected: {wtype}")

    if sheet2.max_row < 3:
        print("[INIT] No outline found. Generating...")
        plan = global_plan(wtype)
        build_outline(sheet2, plan)
        wb.save(workbook_path)
        print("[INIT] Outline saved.")

    sections = load_outline(sheet2)

    if sheet3.max_row == 1:
        sheet3["A1"] = "Prompt"
        sheet3["B1"] = "Response"
        sheet3["C1"] = "SectionIndex"
        sheet3["D1"] = "PromptIndex"
        sheet3["E1"] = "TargetWords"

    print("[START] Autonomous writing begins...")

    for sec in sections:
        if sec["completed"] < sec["prompts"]:
            write_section(wb, sheet3, sheet2, sec, desc, wtype, temp, sections)

    print("[DONE] All sections complete. Saving metadata...")

    sheet5["A1"] = "writing_type"
    sheet5["B1"] = wtype
    sheet5["A2"] = "temperature"
    sheet5["B2"] = temp
    sheet5["A3"] = "description"
    sheet5["B3"] = desc

    wb.save(workbook_path)

    print("[DONE] Project fully written and saved.")'''

# ===== Engine O source =====
#===ENGINE_O_START===
O_CODE = r'''import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

CONTROL_SHEET = "Control"
MAX_TABS = 20

def ensure_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)

def get_column_a_priority_tasks(ws):
    """
    Read A1 downward until first blank.
    Remove duplicates while preserving order (A3 beats A4 overlap).
    """
    ordered = []
    seen = set()
    row = 1
    while True:
        value = ws.cell(row=row, column=1).value
        if value is None or str(value).strip() == "":
            break
        raw = str(value).strip()
        norm = raw.lower()
        if norm not in seen:
            ordered.append(raw)
            seen.add(norm)
        row += 1
    return ordered

def list_tabs_progress(wb):
    tabs = wb.sheetnames[:MAX_TABS]
    print("Tab list:")
    for name in tabs:
        print(f"- {name}")
    print("end of list")
    return tabs

def _has_text_or_value(v):
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() != ""
    return True

def detect_ostensible_borders(ws):
    """
    Border heuristic:
    - Find first row with any text in first 5 active columns
    - Track last populated row, then allow +3 row cushion after blank streak
    - Find last populated column and start writable border after 3 empty columns
    """
    max_row_scan = max(ws.max_row, 1) + 30
    max_col_scan = max(ws.max_column, 1) + 10

    left_text_col = None
    for c in range(1, max_col_scan + 1):
        for r in range(1, min(max_row_scan, 60) + 1):
            if _has_text_or_value(ws.cell(row=r, column=c).value):
                left_text_col = c
                break
        if left_text_col is not None:
            break
    if left_text_col is None:
        return (1, 1, 1, 6)

    probe_cols = list(range(left_text_col, left_text_col + 5))

    top_row = None
    for r in range(1, max_row_scan + 1):
        if any(_has_text_or_value(ws.cell(row=r, column=c).value) for c in probe_cols):
            top_row = r
            break
    if top_row is None:
        top_row = 1

    last_text_row = top_row
    blank_streak = 0
    for r in range(top_row, max_row_scan + 1):
        row_has_text = any(_has_text_or_value(ws.cell(row=r, column=c).value) for c in probe_cols)
        if row_has_text:
            last_text_row = r
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak >= 8:
                break
    bottom_row = last_text_row + 3

    last_text_col = left_text_col
    blank_col_streak = 0
    row_slice_end = max(bottom_row, top_row + 5)
    for c in range(left_text_col, max_col_scan + 1):
        col_has_text = any(
            _has_text_or_value(ws.cell(row=r, column=c).value)
            for r in range(top_row, row_slice_end + 1)
        )
        if col_has_text:
            last_text_col = c
            blank_col_streak = 0
        else:
            blank_col_streak += 1
            if blank_col_streak >= 3:
                break
    entry_start_col = last_text_col + 4
    return (top_row, bottom_row, left_text_col, entry_start_col)

def parse_sheet_instruction(task_text):
    if ":" not in task_text:
        return None, task_text
    left, right = task_text.split(":", 1)
    return left.strip(), right.strip()

def compute_minimal_result(ws, instruction):
    """
    Strict spreadsheet output: return only compact values.
    Supported forms:
    - LITERAL <value>
    - SUM <A>
    - COUNT <A>
    """
    text = instruction.strip()
    if not text:
        return ""

    lit = re.match(r"^LITERAL\s+(.+)$", text, flags=re.IGNORECASE)
    if lit:
        return lit.group(1).strip()

    sum_m = re.match(r"^SUM\s+([A-Z]+)$", text, flags=re.IGNORECASE)
    if sum_m:
        col = column_index_from_string(sum_m.group(1).upper())
        total = 0.0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (int, float)):
                total += v
        return str(int(total) if float(total).is_integer() else total)

    cnt_m = re.match(r"^COUNT\s+([A-Z]+)$", text, flags=re.IGNORECASE)
    if cnt_m:
        col = column_index_from_string(cnt_m.group(1).upper())
        count = 0
        for r in range(1, ws.max_row + 1):
            if _has_text_or_value(ws.cell(row=r, column=col).value):
                count += 1
        return str(count)

    return text.splitlines()[0].strip()

def write_result_in_first_blank(ws, value, top_row, bottom_row, entry_start_col):
    for c in range(entry_start_col, entry_start_col + 20):
        for r in range(top_row, bottom_row + 1):
            cell = ws.cell(row=r, column=c)
            if not _has_text_or_value(cell.value):
                cell.value = value
                return (r, c)
    return None

def run_o_operator(wb):
    tabs = list_tabs_progress(wb)
    if not tabs:
        print("No tabs found")
        return

    task_sheet = wb[tabs[0]]
    tasks = get_column_a_priority_tasks(task_sheet)
    if not tasks:
        print("No priority tasks in first tab column A")
    else:
        print(f"Priority tasks loaded: {len(tasks)}")

    completed_tabs = []
    for tab_name in tabs:
        ws = wb[tab_name]
        top_row, bottom_row, left_col, entry_col = detect_ostensible_borders(ws)
        print(
            f"[TAB] {tab_name} | border rows {top_row}-{bottom_row} "
            f"| left {get_column_letter(left_col)} | entry {get_column_letter(entry_col)}"
        )

        matched_task = None
        for task in tasks:
            target_sheet, instruction = parse_sheet_instruction(task)
            if target_sheet and target_sheet.lower() != tab_name.lower():
                continue
            matched_task = instruction
            break

        if matched_task:
            result = compute_minimal_result(ws, matched_task)
            loc = write_result_in_first_blank(ws, result, top_row, bottom_row, entry_col)
            if loc:
                r, c = loc
                print(f"[WRITE] {tab_name}!{get_column_letter(c)}{r} <- {result}")
            else:
                print(f"[WRITE] {tab_name}: no blank cell in entry border")

        completed_tabs.append(tab_name)
        print(f"[DONE] {tab_name}")

    print(f"Reached last tab: {completed_tabs[-1]}")
    print("Tab completion tracker:")
    for name in completed_tabs:
        print(f"- {name}")
    print("No additional tabs searched.")

def task_new_pool(wb, row):
    target_sheet = row.get("TargetSheet")
    domain = (row.get("Domain") or "").strip().upper()
    notes = row.get("Notes") or ""
    if not target_sheet:
        print("NEW_POOL missing TargetSheet; skipping")
        return
    ws = ensure_sheet(wb, target_sheet)
    if domain == "MARKETING":
        headers = ["ItemID", "Channel", "Audience", "Message", "Cost", "ExpectedReach", "SourceNotes"]
    elif domain == "FINANCE":
        headers = ["EntryID", "Category", "Subcategory", "Amount", "Frequency", "Assumptions", "SourceNotes"]
    elif domain == "CUSTOMER":
        headers = ["ContactID", "Segment", "Need", "PainPoint", "Comment", "Source"]
    else:
        headers = ["ItemID", "Type", "Description", "Value1", "Value2", "Notes"]
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        if notes:
            ws["A2"] = f"# {notes}"

def task_summarize_pool(wb, row):
    source_sheet = row.get("TargetSheet")
    output_sheet = row.get("OutputSheet") or f"{source_sheet}_Summary"
    description = row.get("Description") or ""
    if not source_sheet or source_sheet not in wb.sheetnames:
        print("SUMMARIZE_POOL: source sheet not found; skipping")
        return
    src_df = pd.DataFrame(wb[source_sheet].values)
    if src_df.empty:
        print("SUMMARIZE_POOL: source sheet empty; skipping")
        return
    src_df.columns = src_df.iloc[0]
    src_df = src_df[1:]
    ws_out = ensure_sheet(wb, output_sheet)
    ws_out.delete_rows(1, ws_out.max_row or 1)
    ws_out["A1"] = f"Summary of {source_sheet}"
    ws_out["A2"] = description
    start_row = 4
    ws_out[f"A{start_row}"] = "Column"
    ws_out[f"B{start_row}"] = "Non-empty"
    ws_out[f"C{start_row}"] = "Total"
    ws_out[f"D{start_row}"] = "% Filled"
    for idx, col in enumerate(src_df.columns, start=1):
        col_series = src_df[col]
        non_empty = col_series.notna().sum()
        total = len(col_series)
        pct = (non_empty / total * 100) if total else 0
        row_i = start_row + idx
        ws_out[f"A{row_i}"] = str(col)
        ws_out[f"B{row_i}"] = int(non_empty)
        ws_out[f"C{row_i}"] = int(total)
        ws_out[f"D{row_i}"] = float(round(pct, 1))

def task_decision_matrix(wb, row):
    options_sheet = row.get("TargetSheet")
    output_sheet = row.get("OutputSheet") or f"{options_sheet}_Decision"
    criteria_text = row.get("Criteria") or ""
    goal = row.get("Goal") or ""
    if not options_sheet or options_sheet not in wb.sheetnames:
        print("DECISION_MATRIX: options sheet not found; skipping")
        return
    opt_df = pd.DataFrame(wb[options_sheet].values)
    if opt_df.empty:
        print("DECISION_MATRIX: options sheet empty; skipping")
        return
    opt_df.columns = opt_df.iloc[0]
    opt_df = opt_df[1:]
    if "Option" not in opt_df.columns:
        print("DECISION_MATRIX: 'Option' column missing; skipping")
        return
    criteria = [c.strip() for c in criteria_text.split(",") if c.strip()]
    ws_out = ensure_sheet(wb, output_sheet)
    ws_out.delete_rows(1, ws_out.max_row or 1)
    ws_out["A1"] = "Decision Matrix"
    ws_out["A2"] = f"Goal: {goal}"
    ws_out["A3"] = f"Criteria (1-5 rating, higher is better): {', '.join(criteria)}"
    header_row = 5
    ws_out[f"A{header_row}"] = "Option"
    for i, c in enumerate(criteria, start=1):
        ws_out[f"{get_column_letter(i + 1)}{header_row}"] = c
    total_col = get_column_letter(len(criteria) + 2)
    ws_out[f"{total_col}{header_row}"] = "TotalScore"
    for idx, (_, opt_row) in enumerate(opt_df.iterrows(), start=1):
        row_i = header_row + idx
        ws_out[f"A{row_i}"] = opt_row["Option"]
        for j in range(1, len(criteria) + 1):
            ws_out[f"{get_column_letter(j + 1)}{row_i}"] = None
        if criteria:
            ws_out[f"{total_col}{row_i}"] = f"=SUM(B{row_i}:{get_column_letter(len(criteria) + 1)}{row_i})"

def run_control_tasks_if_present(wb):
    if CONTROL_SHEET not in wb.sheetnames:
        print(f"No '{CONTROL_SHEET}' sheet found; skipping control tasks")
        return
    control_df = pd.DataFrame(wb[CONTROL_SHEET].values)
    if control_df.empty:
        print("Control sheet empty; nothing to do")
        return
    control_df.columns = control_df.iloc[0]
    control_df = control_df[1:]
    task_handlers = {
        "NEW_POOL": task_new_pool,
        "SUMMARIZE_POOL": task_summarize_pool,
        "DECISION_MATRIX": task_decision_matrix,
    }
    for _, row in control_df.iterrows():
        task_type = (row.get("TaskType") or "").strip().upper()
        if not task_type:
            break
        handler = task_handlers.get(task_type)
        if not handler:
            print(f"Unknown TaskType '{task_type}', skipping")
            continue
        print(f"Running task: {task_type}")
        handler(wb, row)

def process_operator(path: Path):
    wb = load_workbook(path)
    run_o_operator(wb)
    run_control_tasks_if_present(wb)
    wb.save(path)
    print(f"Updated workbook saved to {path}")

def main():
    parser = argparse.ArgumentParser(description="O-Operator spreadsheet engine")
    parser.add_argument("workbook", nargs="?", default="Worthy.xlsx", help="Path to .xlsx file")
    args = parser.parse_args()
    path = Path(args.workbook)
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")
    process_operator(path)

if __name__ == "__main__":
    main()'''

# ===== Engine R source =====
#===ENGINE_R_START===
R_CODE = r'''import sys
import os
import shutil
import logging
from dataclasses import dataclass
from typing import List, Tuple, Dict, Optional

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet
from openai import OpenAI

# ------------------ CONFIG ------------------

MODEL_EMPLOYEE = "gpt-4.1-mini"
MODEL_EVALUATOR = "gpt-4.1-mini"
MODEL_EXECUTIVE = "gpt-4.1-mini"

SHEET_GOALS = "Sheet1"          # Mission statement(s)
SHEET_SERIES = "Sheet2"         # Series of tasks (A–M)
SHEET_GOAL_TABLE = "Sheet3"     # 10 goals and best plans/summaries
SHEET_DEPTS = "Sheet4"          # Dynamic departments/tasks

SERIES_BLOCK_SIZE = 20          # 19 tasks + 1 evaluator row
NUM_GOALS = 10                  # Number of rubric goals

MIN_GOAL_GRADE = "C+"           # Per-goal minimum for "good enough"

client = OpenAI()

import sys

def parse_cli_untilgpa():
    """
    Reads --untilgpa <value> from command line if provided.
    Returns float or None.
    """
    if "--untilgpa" not in sys.argv:
        return None
    try:
        idx = sys.argv.index("--untilgpa")
        return float(sys.argv[idx + 1])
    except:
        return None

# ------------------ GRADES ------------------

GRADE_RANK = [
    "F", "E", "D-", "D", "D+",
    "C-", "C", "C+",
    "B-", "B", "B+",
    "A-", "A", "A+"
]
GRADE_INDEX = {g: i for i, g in enumerate(GRADE_RANK)}

GRADE_TO_GPA = {
    "A+": 4.3,
    "A": 4.0,
    "A-": 3.7,
    "B+": 3.3,
    "B": 3.0,
    "B-": 2.7,
    "C+": 2.3,
    "C": 2.0,
    "C-": 1.7,
    "D+": 1.3,
    "D": 1.0,
    "D-": 0.7,
    "E": 0.3,
    "F": 0.0,
}

# ------------------ DATA STRUCTURES ------------------

@dataclass
class TaskDef:
    local_index: int       # 1–19 position within a series block
    department: str
    role: str
    style: str             # PERSONAL / CAREER / SCHOOL
    task_name: str
    code: str
    task_type: str = "AUTO"  # In this version: set by mission mode ("RESEARCH" or "ACTION")

# ------------------ LOGGING & UTILITIES ------------------

def setup_logging(debug: bool):
    level = logging.DEBUG if debug else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler("local_research.log", encoding="utf-8"),
            logging.StreamHandler()
        ]
    )
    logging.info("Logging initialized. Debug=%s", debug)


def emergency_stop_check() -> bool:
    return os.path.exists(".stop")


def backup_workbook(path: str, series_number: int):
    base, ext = os.path.splitext(path)
    backup_path = f"{base}_backup_latest{ext}"
    try:
        shutil.copyfile(path, backup_path)
        logging.info("Backup created for series %d: %s", series_number, backup_path)
    except Exception as e:
        logging.error("Failed to create backup: %s", e)


def call_chat(model: str, system_prompt: str, user_prompt: str) -> str:
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        )
        content = resp.choices[0].message.content.strip()
        return content
    except Exception as e:
        logging.error("OpenAI API error: %s", e)
        raise

# ------------------ MODE DETECTION ------------------

def detect_mission_mode(mission: str) -> str:
    """
    Decide whether the mission is RESEARCH or ACTION.
    - If it clearly asks for a decision/choice/outcome → ACTION
    - Otherwise → RESEARCH (default)
    """
    text = (mission or "").lower()

    action_keywords = [
        "decide", "decision", "choose", "pick", "select",
        "which should i", "which should we", "what should i do",
        "what should we do", "best way to", "best option", "best choice",
        "prioritize", "rank", "optimize", "select the best", "make a plan"
    ]
    for k in action_keywords:
        if k in text:
            logging.info("Mission mode detected: ACTION (keyword: %s)", k)
            return "ACTION"

    logging.info("Mission mode detected: RESEARCH (default informational/exploratory).")
    return "RESEARCH"

# ------------------ SHEET HELPERS ------------------

def find_active_series_block(ws: Worksheet) -> Tuple[int, int, int]:
    """
    Returns (start_row, end_row, series_number) of first incomplete 20-row block.
    A block is COMPLETE if:
      - For rows 1–19: A,B,C and D–M are all non-empty
      - For row 20: A and B are non-empty
    """
    row = 1
    while True:
        start = row
        end = row + SERIES_BLOCK_SIZE - 1
        block_complete = True

        # rows 1–19
        for r in range(start, start + 19):
            if ws.cell(row=r, column=1).value in (None, ""):
                block_complete = False
                break
            if ws.cell(row=r, column=2).value in (None, ""):
                block_complete = False
                break
            if ws.cell(row=r, column=3).value in (None, ""):
                block_complete = False
                break
            for col in range(4, 4 + NUM_GOALS):
                if ws.cell(row=r, column=col).value in (None, ""):
                    block_complete = False
                    break
            if not block_complete:
                break

        # Row 20 A/B
        if ws.cell(row=end, column=1).value in (None, ""):
            block_complete = False
        if ws.cell(row=end, column=2).value in (None, ""):
            block_complete = False

        if not block_complete:
            series_num = ((start - 1) // SERIES_BLOCK_SIZE) + 1
            return start, end, series_num

        row += SERIES_BLOCK_SIZE


def get_latest_goal(ws: Worksheet) -> Optional[str]:
    latest = None
    for c in ws["A"]:
        if c.value is None or str(c.value).strip() == "":
            break
        latest = str(c.value)
    return latest

# ------------------ SHEET3 GOAL TABLE ------------------

def ensure_goal_table_sheet(wb):
    if SHEET_GOAL_TABLE in wb.sheetnames:
        ws = wb[SHEET_GOAL_TABLE]
    else:
        ws = wb.create_sheet(SHEET_GOAL_TABLE)

    ws.cell(row=1, column=1).value = "Goal"
    ws.cell(row=1, column=2).value = "Grade"
    ws.cell(row=1, column=3).value = "Line found"
    ws.cell(row=1, column=4).value = "Base Employee #"
    ws.cell(row=1, column=5).value = "Department"
    ws.cell(row=1, column=6).value = "Grade (Current)"
    ws.cell(row=1, column=7).value = "Current Plan/Summary"
    return ws


def parse_goal_weight(goal_text: str) -> float:
    if not goal_text:
        return 0.0
    text = goal_text.lower()
    idx = text.find("weight=")
    if idx == -1:
        return 0.0
    after = text[idx + len("weight="):]
    num = ""
    for ch in after:
        if ch in "0123456789.":
            num += ch
        elif num:
            break
    try:
        return float(num)
    except Exception:
        return 0.0


def get_goal_weights(ws_goal_table: Worksheet) -> List[float]:
    weights = []
    for i in range(NUM_GOALS):
        row = 2 + i
        goal_text = ws_goal_table.cell(row=row, column=1).value
        if goal_text in (None, ""):
            weights.append(0.0)
        else:
            w = parse_goal_weight(str(goal_text))
            weights.append(max(w, 0.0))
    total = sum(weights)
    if total <= 0:
        n = max(1, len(weights))
        return [1.0 / n] * n
    return [w / total for w in weights]


def initialize_goals_if_needed(ws_goals: Worksheet, ws_goal_table: Worksheet):
    has_goals = any(
        ws_goal_table.cell(row=r, column=1).value not in (None, "")
        for r in range(2, ws_goal_table.max_row + 1)
    )
    if has_goals:
        return

    mission = get_latest_goal(ws_goals)
    if not mission:
        return

    logging.info("Initializing 10 goals in Sheet3 based on mission: %s", mission)

    system_prompt = (
        "You are an executive designing a grading rubric for work addressing a single mission.\n\n"
        "You must create EXACTLY 10 goals that will be used to evaluate candidate outputs.\n"
        "- 2–3 of these goals should be PRIMARY and together should cover ~80% of the core concern.\n"
        "- The remaining 7–8 should be SECONDARY, covering edge cases, risks, or special constraints, "
        "sharing ~20%.\n\n"
        "Each goal must:\n"
        "- Be 4–6 sentences long, very specific and constraint-rich.\n"
        "- Include practical constraints (clarity, focus, usefulness, feasibility, impact, etc.).\n\n"
        "Output exactly 10 lines, each:\n"
        "PRIMARY | weight=<0-1> | <4-6 sentence goal text>\n"
        "or\n"
        "SECONDARY | weight=<0-1> | <4-6 sentence goal text>\n\n"
        "Weights must be positive and sum to about 1.0 overall. No extra commentary."
    )
    user_prompt = f"The mission is:\n\"{mission}\""

    raw = call_chat(MODEL_EXECUTIVE, system_prompt, user_prompt)
    lines = [l.strip() for l in raw.splitlines() if l.strip()]
    lines = lines[:NUM_GOALS]

    parsed = []
    for line in lines:
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 3:
            continue
        primary_flag = parts[0].upper()
        weight_text = parts[1]
        goal_text = "|".join(parts[2:]).strip()
        w = parse_goal_weight(weight_text)
        if w <= 0:
            w = 0.1
        label = "PRIMARY" if "PRIMARY" in primary_flag else "SECONDARY"
        parsed.append((label, w, goal_text))

    while len(parsed) < NUM_GOALS:
        parsed.append(("SECONDARY", 0.1, "Backup goal for mission alignment."))

    total = sum(w for _, w, _ in parsed)
    if total <= 0:
        total = 1.0
    norm = [(label, w / total, text) for (label, w, text) in parsed]

    for i in range(NUM_GOALS):
        row = 2 + i
        label, w, text = norm[i]
        goal_str = f"{label} | weight={w:.3f} | {text}"
        ws_goal_table.cell(row=row, column=1).value = goal_str
        ws_goal_table.cell(row=row, column=2).value = ""
        ws_goal_table.cell(row=row, column=3).value = ""
        ws_goal_table.cell(row=row, column=4).value = ""
        ws_goal_table.cell(row=row, column=5).value = ""
        ws_goal_table.cell(row=row, column=6).value = ""
        ws_goal_table.cell(row=row, column=7).value = ""

    logging.info("Initialized 10 goals in Sheet3.")

# ------------------ SHEET4: DEPARTMENTS & TASKS ------------------

def ensure_dept_sheet(wb):
    if SHEET_DEPTS in wb.sheetnames:
        ws = wb[SHEET_DEPTS]
    else:
        ws = wb.create_sheet(SHEET_DEPTS)
    ws.cell(row=1, column=1).value = "TaskIndex"
    ws.cell(row=1, column=2).value = "Department"
    ws.cell(row=1, column=3).value = "DepartmentRole"
    ws.cell(row=1, column=4).value = "Style"
    ws.cell(row=1, column=5).value = "TaskName"
    ws.cell(row=1, column=6).value = "TaskCode"
    ws.cell(row=1, column=7).value = "TaskType"  # For visibility; we set based on mode
    return ws


def generate_departments_and_tasks_if_needed(ws_goals: Worksheet, ws_depts: Worksheet):
    has_tasks = any(
        ws_depts.cell(row=r, column=1).value not in (None, "")
        for r in range(2, ws_depts.max_row + 1)
    )
    if has_tasks:
        return

    mission = get_latest_goal(ws_goals)
    if not mission:
        return

    logging.info("Generating dynamic departments and tasks based on mission: %s", mission)

    system_prompt = (
        "You are an executive architect designing 5 departments and 19 tasks to analyze and advance a mission.\n\n"
        "First, infer whether the mission sounds primarily PERSONAL, CAREER, or SCHOOL. "
        "Then design departments and tasks appropriate to that world:\n"
        "- If PERSONAL: think DIY, at-home, individual or family-scale.\n"
        "- If CAREER: think early or mid-career professional.\n"
        "- If SCHOOL: think coursework or projects with limited time/resources.\n\n"
        "You must:\n"
        "- Create exactly 5 departments.\n"
        "- Assign them a STYLE: PERSONAL, CAREER, or SCHOOL.\n"
        "- Create exactly 19 tasks across these departments with a distribution like 3,3,3,5,5.\n"
        "- Each task must be concrete and clearly tied to the mission.\n\n"
        "Output EXACTLY 19 lines, one per task, in this format:\n"
        "TASK <index_1_to_19> | Department=<dept name> | Style=<PERSONAL/Career/School> | "
        "Role=<short dept role> | Task=<specific task objective>\n\n"
        "Do NOT include task types; the system will set those based on the mission mode.\n"
        "No extra commentary."
    )

    user_prompt = f"The mission is:\n\"{mission}\""

    raw = call_chat(MODEL_EXECUTIVE, system_prompt, user_prompt)
    lines = [l.strip() for l in raw.splitlines() if l.strip()]
    lines = [ln for ln in lines if ln.upper().startswith("TASK")]
    lines = lines[:19]

    if not lines:
        logging.error("No parsable TASK lines returned for departments.")
        return

    for line in lines:
        try:
            parts = [p.strip() for p in line.split("|")]
            if len(parts) < 5:
                continue

            first = parts[0]
            _, idx_str = first.split(maxsplit=1)
            local_index = int(idx_str.strip())

            dept_part = parts[1]
            style_part = parts[2]
            role_part = parts[3]
            task_part = "|".join(parts[4:])

            dept_name = dept_part.split("=", 1)[1].strip()
            style_raw = style_part.split("=", 1)[1].strip().upper()
            if "PERSONAL" in style_raw:
                style = "PERSONAL"
            elif "SCHOOL" in style_raw:
                style = "SCHOOL"
            else:
                style = "CAREER"
            role = role_part.split("=", 1)[1].strip()
            task_name = task_part.split("=", 1)[1].strip()

            if local_index < 1 or local_index > 19:
                continue

            row = 1 + local_index
            ws_depts.cell(row=row, column=1).value = local_index
            ws_depts.cell(row=row, column=2).value = dept_name
            ws_depts.cell(row=row, column=3).value = role
            ws_depts.cell(row=row, column=4).value = style
            ws_depts.cell(row=row, column=5).value = task_name
            ws_depts.cell(row=row, column=6).value = f"T{local_index:02d}"
            ws_depts.cell(row=row, column=7).value = ""  # Will be set to RESEARCH/ACTION based on mission mode
        except Exception as e:
            logging.error("Error parsing dept line '%s': %s", line, e)

    logging.info("Dynamic departments and tasks created in Sheet4.")


def load_tasks_from_sheet4(ws_depts: Worksheet) -> List[TaskDef]:
    tasks: List[TaskDef] = []
    for r in range(2, ws_depts.max_row + 1):
        idx_val = ws_depts.cell(row=r, column=1).value
        if idx_val in (None, ""):
            continue
        try:
            local_index = int(idx_val)
        except Exception:
            continue

        dept = ws_depts.cell(row=r, column=2).value or "Department"
        role = ws_depts.cell(row=r, column=3).value or ""
        style = (ws_depts.cell(row=r, column=4).value or "CAREER").upper()
        task_name = ws_depts.cell(row=r, column=5).value or "Task objective"
        code = ws_depts.cell(row=r, column=6).value or f"T{local_index:02d}"
        task_type_raw = ws_depts.cell(row=r, column=7).value or "AUTO"

        tasks.append(TaskDef(
            local_index=local_index,
            department=str(dept),
            role=str(role),
            style=style,
            task_name=str(task_name),
            code=str(code),
            task_type=str(task_type_raw).upper()
        ))
    tasks.sort(key=lambda t: t.local_index)
    return tasks


def build_department_maps(tasks: List[TaskDef]) -> Tuple[List[str], Dict[str, int], Dict[int, str]]:
    departments: List[str] = []
    dept_last_index: Dict[str, int] = {}
    local_to_dept: Dict[int, str] = {}
    for t in tasks:
        if t.department not in departments:
            departments.append(t.department)
        dept_last_index[t.department] = max(dept_last_index.get(t.department, 0), t.local_index)
        local_to_dept[t.local_index] = t.department
    return departments, dept_last_index, local_to_dept

# ------------------ HISTORY HELPERS ------------------

def collect_department_history_same_series(ws: Worksheet, task: TaskDef, tasks: List[TaskDef], start_row: int) -> List[str]:
    hist = []
    for t in tasks:
        if t.local_index >= task.local_index:
            continue
        if t.department != task.department:
            continue
        abs_row = start_row + (t.local_index - 1)
        val = ws.cell(row=abs_row, column=1).value
        if val not in (None, ""):
            hist.append(str(val))
    return hist


def collect_task_history_prior_series(ws: Worksheet, local_index: int, current_series_num: int) -> List[str]:
    hist = []
    for s in range(1, current_series_num):
        row = (s - 1) * SERIES_BLOCK_SIZE + local_index
        val = ws.cell(row=row, column=1).value
        if val not in (None, ""):
            hist.append(str(val))
    return hist

# ------------------ COLUMN A ------------------

def ensure_column_a(ws_series: Worksheet, mission: str, tasks: List[TaskDef],
                    departments: List[str], dept_last_index: Dict[str, int],
                    start_row: int, end_row: int, series_num: int):
    block_offset = start_row - 1
    logging.info("Filling Column A for series #%d (rows %d-%d)", series_num, start_row, end_row)

    for task in tasks:
        abs_row = block_offset + task.local_index
        cell = ws_series.cell(row=abs_row, column=1)
        if cell.value not in (None, ""):
            continue

        same_series_hist = collect_department_history_same_series(ws_series, task, tasks, start_row)
        prior_series_hist = collect_task_history_prior_series(ws_series, task.local_index, series_num)

        system_prompt = (
            f"You are a {task.department} producing a single question or objective.\n"
            f"The mission is:\n\"{mission}\"\n\n"
            f"Your high-level role: {task.role}\n"
            f"Your specific task: {task.task_name}\n\n"
            "You only see your own department's thinking. "
            "You do not see other departments.\n"
            "Do NOT optimize for evaluator goals, grading criteria, or scores.\n"
            "Context style: PERSONAL, CAREER, or SCHOOL; keep tone natural and non-academic."
        )

        history_text = ""
        if prior_series_hist:
            history_text += "Previous questions for this task across earlier series:\n"
            for q in prior_series_hist:
                history_text += f"- {q}\n"
        if same_series_hist:
            history_text += "\nEarlier entries from your department in this series:\n"
            for q in same_series_hist:
                history_text += f"- {q}\n"

        user_prompt = (
            history_text
            + "\n\nGenerate a completely NEW question or objective for this task. "
              "Do NOT refine or rephrase earlier questions. "
              "Avoid repeating earlier content. "
              "Do NOT attempt to optimize for evaluator goals, grading criteria, or scores. "
              "Keep the question concrete, practical, and aligned only with your department's role and the mission. "
              "Use a natural, non-academic tone."
        )

        logging.info("Generating A-row %d for %s (%s), series %d",
                     abs_row, task.department, task.code, series_num)
        result = call_chat(MODEL_EMPLOYEE, system_prompt, user_prompt)
        cell.value = result

        if emergency_stop_check():
            logging.warning(".stop detected after writing A-row %d.", abs_row)
            return

    # Evaluator summary at row 20, col A
    eval_cell = ws_series.cell(row=end_row, column=1)
    if eval_cell.value in (None, ""):
        dept_prompts = []
        for dept in departments:
            last_idx = dept_last_index[dept]
            abs_row = block_offset + last_idx
            val = ws_series.cell(row=abs_row, column=1).value
            if val not in (None, ""):
                dept_prompts.append((dept, str(val)))

        if dept_prompts:
            system_prompt = (
                "You are the Evaluator. Read the final prompt from each department and create one unified summary. "
                "Do not list departments or label sections. Blend the ideas into a coherent narrative in natural, "
                "accessible language. Avoid technical jargon; do not sound academic. Do not create new missions "
                "or recommendations; summarize only. ~150 words, max 10 sentences."
            )
            prompt_lines = [f"{d}: {t}" for (d, t) in dept_prompts]
            user_prompt = (
                f"Original mission:\n{mission}\n\n"
                "Final department prompts:\n\n" +
                "\n".join(prompt_lines) +
                "\n\nWrite a unified summary (~150 words, max 10 sentences):"
            )
            logging.info("Generating Evaluator prompt summary at row %d (A).", end_row)
            result = call_chat(MODEL_EVALUATOR, system_prompt, user_prompt)
            eval_cell.value = result

# ------------------ DYNAMIC GOAL FOCUS ------------------

def get_primary_goal_index_from_grades(grade_cells: List[Optional[str]]) -> Optional[int]:
    """
    Given a list of grade strings for G1..G10, return the index (0-based) of the highest grade.
    If no valid grades, return None.
    """
    best_idx = None
    best_rank = -1
    for i, g in enumerate(grade_cells):
        if g in (None, ""):
            continue
        gs = str(g).strip()
        if gs not in GRADE_INDEX:
            continue
        rank = GRADE_INDEX[gs]
        if rank > best_rank:
            best_rank = rank
            best_idx = i
    return best_idx

# ------------------ COLUMN B (MISSION-AWARE + GOAL FOCUS) ------------------

def ensure_column_b(ws_series: Worksheet, mission: str, tasks: List[TaskDef],
                    departments: List[str], dept_last_index: Dict[str, int],
                    ws_goal_table: Worksheet,
                    start_row: int, end_row: int, series_num: int):
    block_offset = start_row - 1
    logging.info("Filling Column B for series #%d (rows %d-%d)", series_num, start_row, end_row)

    for task in tasks:
        abs_row = block_offset + task.local_index
        q_cell = ws_series.cell(row=abs_row, column=1)
        a_cell = ws_series.cell(row=abs_row, column=2)

        if q_cell.value in (None, ""):
            continue
        if a_cell.value not in (None, ""):
            continue

        question = str(q_cell.value)

        # Dynamic goal focus: look at previous series' grades for this row
        primary_goal_text = None
        primary_goal_idx = None

        if series_num > 1:
            prev_series = series_num - 1
            prev_row = (prev_series - 1) * SERIES_BLOCK_SIZE + task.local_index
            # Ensure prev_row exists and has some grades
            grade_cells = []
            for goal_col_offset in range(NUM_GOALS):  # D..M
                col = 4 + goal_col_offset
                grade_cells.append(ws_series.cell(row=prev_row, column=col).value)

            pg_idx = get_primary_goal_index_from_grades(grade_cells)
            if pg_idx is not None:
                primary_goal_idx = pg_idx
                # Goals in Sheet3: row = 2 + idx, column 1
                goal_row = 2 + pg_idx
                gtext = ws_goal_table.cell(row=goal_row, column=1).value
                if gtext not in (None, ""):
                    primary_goal_text = str(gtext)

        # Mission-aware, drift-resistant system prompt
        system_prompt = (
            f"You are responding ONLY from the perspective of the {task.department}.\n\n"
            "Your job is to provide clear, factual, mission-relevant information.\n\n"
            f"The overall mission is:\n\"{mission}\"\n\n"
            f"Context style: {task.style} (PERSONAL, CAREER, or SCHOOL).\n"
            f"Role: {task.role}\n"
            f"Internal task: {task.task_name}.\n\n"
            "Rules:\n"
            "1. Your answer MUST stay directly connected to the mission at all times.\n"
            "2. If the question in Column A starts drifting away from the mission, reinterpret it in a way that serves the mission and answer that version.\n"
            "3. Do not introduce unrelated topics, long-term plans, studies, or recommendations.\n"
            "4. Focus on the most meaningful information that helps understand the mission.\n"
            "5. Be concise, clear, and factual. Do not paraphrase the question.\n"
            "6. In research mode, do NOT give advice, decisions, or next steps.\n"
        )

        user_prompt = f"Question/objective in Column A:\n\"{question}\"\n"

        if primary_goal_idx is not None and primary_goal_text:
            user_prompt += (
                f"\nFor this row, give extra attention to Goal #{primary_goal_idx + 1}:\n"
                f"\"{primary_goal_text}\"\n"
                "You do not need to mention the goal explicitly, but prioritize information that helps evaluate progress on that goal.\n"
            )

        user_prompt += (
            "\nNow respond with a clear, mission-aligned informational answer. "
            "If the question is awkward or partially off-mission, reshape it mentally and answer the version that best serves the mission."
        )

        logging.info("Generating B-row %d for %s (%s)", abs_row, task.department, task.code)
        result = call_chat(MODEL_EMPLOYEE, system_prompt, user_prompt)
        a_cell.value = result

        if emergency_stop_check():
            logging.warning(".stop detected after writing B-row %d.", abs_row)
            return

    # Evaluator summary at row 20, col B
    eval_cell = ws_series.cell(row=end_row, column=2)
    if eval_cell.value in (None, ""):
        dept_answers = []
        for dept in departments:
            last_idx = dept_last_index[dept]
            abs_row = block_offset + last_idx
            val = ws_series.cell(row=abs_row, column=2).value
            if val not in (None, ""):
                dept_answers.append((dept, str(val)))

        if dept_answers:
            system_prompt = (
                "You are the Evaluator. Create a seamless, integrated summary of final findings from each department. "
                "Do not list departments or sound blocky. Blend findings into one cohesive explanation in clear language, "
                "avoiding jargon or academic tone. Respect depth but stay accessible. ~150 words, max 10 sentences. "
                "Do NOT introduce new conclusions or change the mission."
            )
            answer_lines = [f"{d}: {t}" for (d, t) in dept_answers]
            user_prompt = (
                f"Original mission:\n{mission}\n\n"
                "Final departmental findings:\n\n" +
                "\n".join(answer_lines) +
                "\n\nWrite a unified summary (~150 words, max 10 sentences):"
            )
            logging.info("Generating Evaluator findings summary at row %d (B).", end_row)
            result = call_chat(MODEL_EVALUATOR, system_prompt, user_prompt)
            eval_cell.value = result

# ------------------ COLUMN C SUMMARIZERS ------------------

def summarize_research(answer_text: str, mission: str, goals: List[Tuple[int, str]]) -> str:
    goals_str = ""
    for gid, gtext in goals:
        goals_str += f"Goal {gid}:\n{gtext}\n\n"

    system_prompt = (
        "You rewrite longer answers into concise research summaries.\n"
        "Your job is to summarize WHAT IS KNOWN or WHAT INFORMATION MATTERS, not to advise or decide.\n\n"
        "Mission-awareness:\n"
        "- Focus on information that is directly relevant to the mission.\n"
        "- If parts of the answer drift away from the mission or goals, omit or downplay them.\n\n"
        "Goal-awareness (soft):\n"
        "- Emphasize content that would be helpful when evaluating progress against the mission and its goals.\n\n"
        "Rules:\n"
        "- 1–3 sentences.\n"
        "- Do NOT tell the reader what they should do.\n"
        "- Do NOT propose plans, steps, projects, or experiments.\n"
        "- Do NOT recommend options or make decisions.\n"
        "- Do NOT simply paraphrase or restate the question from Column A.\n"
        "- Focus only on summarizing information, patterns, or key ideas."
    )

    user_prompt = (
        f"Mission:\n{mission}\n\n"
        "Goals (for context, do NOT mention them explicitly):\n"
        f"{goals_str}\n"
        "Original answer:\n"
        f"{answer_text}\n\n"
        "Summarize the key information in 1–3 sentences, without giving advice, recommendations, plans, "
        "or just paraphrasing the original question."
    )
    return call_chat(MODEL_EVALUATOR, system_prompt, user_prompt)


def summarize_action(answer_text: str) -> str:
    system_prompt = (
        "You rewrite longer analytical answers into ONE realistic next action.\n"
        "Rules:\n"
        "- Output exactly ONE clear next action the person should take.\n"
        "- 1–2 sentences max.\n"
        "- Do NOT list multiple options.\n"
        "- Do NOT say 'you could do A or B'.\n"
        "- Do NOT outline a process without a clear next action.\n"
        "- Do NOT say 'do more research' as the main action.\n"
        "- The action must be concrete and feasible for an individual."
    )
    user_prompt = (
        "Original answer:\n"
        f"{answer_text}\n\n"
        "Rewrite this as 1–2 sentences that describe exactly ONE concrete, realistic next action "
        "the person should take next. Do not give multiple options or a decision tree."
    )
    return call_chat(MODEL_EVALUATOR, system_prompt, user_prompt)

# ------------------ GRADING ------------------

# ------------------ SUMMARIZATION + GRADING (FULL REWRITE) ------------------

def summarize_research(answer_text: str, mission: str, goals: List[Tuple[int, str]]) -> str:
    """
    Improved summarizer that:
    - Never removes numbers (to avoid conflicts with numerical-grade requirements)
    - Produces 1–3 sentence summaries
    - Encourages broad view, main point, contrast, and context
    - Removes ALL advice, recommendations, future steps, or 'should' statements
    """

    goals_str = ""
    for gid, gtext in goals:
        goals_str += f"Goal {gid}:\n{gtext}\n\n"

    system_prompt = (
        "You rewrite analytical text into a concise, high-value research summary.\n\n"
        "Rules:\n"
        "- Length: 1–3 sentences.\n"
        "- Preserve all NUMBERS, statistics, quantities, or percentages.\n"
        "- Include (when possible):\n"
        "  1. Broad view (macro framing)\n"
        "  2. Main point (core finding)\n"
        "  3. Contrast/competition (compare factors/trends)\n"
        "  4. Context or implication\n"
        "- Remove ALL advice, recommendations, judgments, or future directives.\n"
        "- Absolutely no 'should', 'need to', 'best option', 'recommend', etc.\n"
        "- Do NOT paraphrase or restate the question.\n"
        "- Focus strictly on information relevant to the mission and factual insights."
    )

    user_prompt = (
        f"Mission:\n{mission}\n\n"
        f"Goals (context only):\n{goals_str}\n"
        f"Original answer:\n{answer_text}\n\n"
        "Produce a 1–3 sentence research summary preserving all numerical information."
    )

    return call_chat(MODEL_EVALUATOR, system_prompt, user_prompt)


def summarize_action(answer_text: str) -> str:
    """
    (UNCHANGED)
    Produces exactly one next-step action.
    """
    system_prompt = (
        "You rewrite analytical content into ONE realistic next action.\n"
        "Rules:\n"
        "- Output exactly ONE clear next action.\n"
        "- 1–2 sentences max.\n"
        "- No lists, no choices, no options.\n"
        "- No recommending 'research more'.\n"
        "- Must be doable by a person immediately."
    )

    user_prompt = (
        f"Original answer:\n{answer_text}\n\n"
        "Rewrite this as exactly one concrete next action a person can take immediately."
    )

    return call_chat(MODEL_EVALUATOR, system_prompt, user_prompt)


def grade_research_output(summary_text: str, goals: List[Tuple[int, str]]) -> Dict[int, str]:
    """
    Upgraded RESEARCH grading engine implementing 7 rules:

    1. Directly answers the goal → +1 increment
    2. Prescriptive language ('should', 'need to', etc.) → -1 increment
       If strongly prescriptive → drop to C or below
       If fully prescriptive → F
    3. If numbers/statistics required but missing → AUTOMATIC F
    4. <=3 sentences → +1 increment (max B+)
    5. >5 sentences OR >75 words → AUTOMATIC F
    6. Structural requirements:
       Must include (broad view, main point, contrast, context)
       Missing 1 → -1
       Missing 2+ → max C
    7. Insight required for A-range (otherwise max B+)
    """

    # ---- STEP 1: Base grades from your rare-A evaluator ----

    base_system_prompt = (
        "You are a highly selective evaluator of research summaries.\n"
        "A-range requires genuine insight.\n"
        "Without real insight, MAX is B+.\n"
        "To score B+ or higher, summary must clearly support the goal.\n\n"
        "Return EXACTLY 10 grades:\n"
        "G1: <grade>\nG2: <grade>\n...\nG10: <grade>\n"
        "Allowed: A+, A, A-, B+, B, B-, C+, C, C-, D+, D, D-, E, F"
    )

    # Build full goal text
    goals_str = ""
    for gid, gtext in goals:
        goals_str += f"Goal {gid}:\n{gtext}\n\n"

    base_user_prompt = (
        f"SUMMARY TO GRADE:\n{summary_text}\n\n"
        f"GOALS:\n{goals_str}\n"
        "Provide only the 10 grades."
    )

    raw = call_chat(MODEL_EVALUATOR, base_system_prompt, base_user_prompt)
    lines = [l.strip() for l in raw.splitlines() if l.strip()]

    base_grades = {}
    for ln in lines:
        if ":" not in ln:
            continue
        left, right = ln.split(":", 1)
        try:
            gnum = int(left.strip()[1:])
        except:
            continue
        grade = right.strip()
        if grade in GRADE_INDEX:
            base_grades[gnum] = grade

    # ---- STEP 2: Analyze summary characteristics ----

    words = summary_text.split()
    word_count = len(words)

    sentences = [s.strip() for s in summary_text.replace("\n", " ").split(".") if s.strip()]
    num_sent = len(sentences)

    # Automatic fail rules
    if num_sent > 5 or word_count > 75:
        return {gid: "F" for gid, _ in goals}

    # Detect prescriptive language
    prescriptive_terms = [
        "should", "need to", "ought", "must",
        "recommend", "suggest", "could consider", "best approach"
    ]
    prescriptive = any(p in summary_text.lower() for p in prescriptive_terms)

    # ---- STEP 3: Structural evaluation ----
    structure_prompt = (
        "Determine presence (1) or absence (0) of the following four elements:\n"
        "1. Broad view\n"
        "2. Main point\n"
        "3. Contrast/competition\n"
        "4. Context/implication\n\n"
        "Output exactly four numbers (0 or 1) separated by spaces.\n\n"
        f"SUMMARY:\n{summary_text}"
    )
    struct_raw = call_chat(MODEL_EVALUATOR, "You evaluate structural presence.", structure_prompt)
    flags = [int(x) for x in struct_raw.split() if x in ["0", "1"]][:4]
    missing = 4 - sum(flags)

    # ---- STEP 4: Apply advanced scoring per goal ----

    final = {}

    for (gid, goal_text) in goals:
        grade = base_grades.get(gid, "C")
        rank = GRADE_INDEX.get(grade, GRADE_INDEX["C"])

        # 3. If numbers required but missing → automatic F
        numbers_required = any(
            key in goal_text.lower()
            for key in ["how many", "rate", "percent", "percentage", "quantify", "statistic", "compare numbers"]
        )
        has_numbers = any(ch.isdigit() for ch in summary_text)

        if numbers_required and not has_numbers:
            final[gid] = "F"
            continue

        # 1. Direct-answer detection → +1 increment
        direct_prompt = (
            "Does the summary directly answer the goal? Respond YES or NO.\n\n"
            f"Goal:\n{goal_text}\n\n"
            f"Summary:\n{summary_text}"
        )
        direct_raw = call_chat(MODEL_EVALUATOR, "Directness evaluator.", direct_prompt)
        if "yes" in direct_raw.lower():
            rank += 1

        # 2. Prescriptive penalty → -1
        if prescriptive:
            rank -= 1

        # 4. Short summaries (<3 sentences) → +1 (max B+)
        if num_sent <= 3 and rank < GRADE_INDEX["B+"]:
            rank += 1

        # 6. Structural penalties
        if missing == 1:
            rank -= 1
        elif missing >= 2:
            rank = min(rank, GRADE_INDEX["C"])

        # 7. Insight gating for A-range
        if rank > GRADE_INDEX["B+"]:
            insight_prompt = (
                "Does the summary contain meaningful, non–obvious insight? Respond YES or NO.\n\n"
                f"{summary_text}"
            )
            insight_raw = call_chat(MODEL_EVALUATOR, "Insight evaluator.", insight_prompt)
            if "no" in insight_raw.lower():
                rank = GRADE_INDEX["B+"]

        # Clamp grade index
        rank = max(0, min(rank, len(GRADE_RANK) - 1))
        final[gid] = GRADE_RANK[rank]

    return final


def grade_action_output(plan_text: str, goals: List[Tuple[int, str]]) -> Dict[int, str]:
    """
    (UNCHANGED)
    Grades ACTION mode where a plan must give exactly one clear next action.
    """
    system_prompt = (
        "You are a strict evaluator grading a next-step action plan.\n\n"
        "Rules:\n"
        "- Must give ONE clear next action.\n"
        "- Multiple options = automatic F.\n"
        "- A plan without a concrete action = automatic F.\n"
        "- Return EXACTLY:\n"
        "G1: <grade>\n...\nG10: <grade>"
    )

    goals_str = ""
    for gid, gtext in goals:
        goals_str += f"Goal {gid}:\n{gtext}\n\n"

    user_prompt = (
        f"Next-step plan:\n{plan_text}\n\n"
        f"Goals:\n{goals_str}\n"
        "Provide 10 grades as instructed."
    )

    raw = call_chat(MODEL_EVALUATOR, system_prompt, user_prompt)
    lines = [l.strip() for l in raw.splitlines() if l.strip()]

    parsed = {}
    for ln in lines:
        if ":" not in ln:
            continue
        left, right = ln.split(":", 1)
        try:
            gid = int(left.strip()[1:])
        except:
            continue
        grade = right.strip()
        if grade in GRADE_INDEX:
            parsed[gid] = grade

    return parsed

# ------------------ COLUMN C + GRADES D–M ------------------

def ensure_outputs_and_grades(ws_series: Worksheet, ws_goal_table: Worksheet, ws_depts: Worksheet,
                              tasks: List[TaskDef],
                              start_row: int, end_row: int,
                              series_num: int, mode: str, mission: str):
    """
    Fill Column C (research summary or action next-step) and D–M (grades)
    for all 19 rows, according to mission mode.
    """
    logging.info("Filling Column C and D–M for series #%d in mode=%s.", series_num, mode)

    # Load goals from Sheet3
    goals: List[Tuple[int, str]] = []
    for i in range(NUM_GOALS):
        row = 2 + i
        val = ws_goal_table.cell(row=row, column=1).value
        if val in (None, ""):
            continue
        goals.append((i + 1, str(val)))
    if not goals:
        logging.warning("No goals in Sheet3; skipping grading.")
        return

    is_research = (mode == "RESEARCH")

    for offset in range(0, 19):
        r = start_row + offset
        q_cell = ws_series.cell(row=r, column=1)
        ans_cell = ws_series.cell(row=r, column=2)
        out_cell = ws_series.cell(row=r, column=3)

        if ans_cell.value in (None, ""):
            continue

        answer_text = str(ans_cell.value)
        question_text = str(q_cell.value) if q_cell.value not in (None, "") else ""

        local_index = (r - start_row) + 1
        task_for_row: Optional[TaskDef] = None
        for t in tasks:
            if t.local_index == local_index:
                task_for_row = t
                break
        if task_for_row is None:
            logging.warning("No TaskDef found for row %d (local_index %d). Skipping.", r, local_index)
            continue

        # Force task_type to mission mode and persist in Sheet4
        forced_type = "RESEARCH" if is_research else "ACTION"
        task_for_row.task_type = forced_type
        sheet4_row = 1 + local_index
        try:
            ws_depts.cell(row=sheet4_row, column=7).value = forced_type
        except Exception:
            pass

        # Column C: summarize based on mode
        if out_cell.value in (None, ""):
            if is_research:
                logging.info("Summarizing row %d as RESEARCH.", r)
                out_text = summarize_research(answer_text, mission, goals)
            else:
                logging.info("Summarizing row %d as ACTION next-step.", r)
                out_text = summarize_action(answer_text)
            out_cell.value = out_text
        else:
            out_text = str(out_cell.value)

        # D–M: grade, if any empty
        need_grades = any(
            ws_series.cell(row=r, column=3 + i + 1).value in (None, "")
            for i in range(NUM_GOALS)
        )
        if not need_grades:
            continue

        if is_research:
            logging.info("Grading RESEARCH summary at row %d.", r)
            parsed_grades = grade_research_output(out_text, goals)
        else:
            logging.info("Grading ACTION plan at row %d.", r)
            parsed_grades = grade_action_output(out_text, goals)

        for i in range(1, NUM_GOALS + 1):
            col = 3 + i  # D=4
            existing = ws_series.cell(row=r, column=col).value
            if existing not in (None, ""):
                continue
            grade = parsed_grades.get(i, None)
            if grade is None or grade not in GRADE_INDEX:
                grade = "C"
            ws_series.cell(row=r, column=col).value = grade

# ------------------ BEST PLAN / SUMMARY TRACKING ------------------

def is_better_grade(new_grade: str, old_grade: Optional[str]) -> bool:
    if new_grade not in GRADE_INDEX:
        return False
    if not old_grade or old_grade not in GRADE_INDEX:
        return True
    return GRADE_INDEX[new_grade] > GRADE_INDEX[old_grade]


def update_best_rows(ws_series: Worksheet, ws_goal_table: Worksheet,
                     local_to_dept: Dict[int, str],
                     start_row: int, end_row: int, series_num: int):
    """
    Update Sheet3 best row per goal using Column C text and D–M grades,
    regardless of mode. Interpreted as "best plan or best summary".
    """
    logging.info("Updating best rows in Sheet3 from series #%d.", series_num)

    for goal_idx in range(NUM_GOALS):
        goal_row = 2 + goal_idx
        best_grade_cell = ws_goal_table.cell(row=goal_row, column=2)
        best_grade = best_grade_cell.value if best_grade_cell.value not in (None, "") else None

        for offset in range(0, 19):
            r = start_row + offset
            grade_cell = ws_series.cell(row=r, column=4 + goal_idx)
            out_cell = ws_series.cell(row=r, column=3)
            if grade_cell.value in (None, "") or out_cell.value in (None, ""):
                continue
            new_grade = str(grade_cell.value).strip()
            if new_grade not in GRADE_INDEX:
                continue

            if not is_better_grade(new_grade, best_grade):
                continue

            best_grade = new_grade
            best_grade_cell.value = new_grade
            ws_goal_table.cell(row=goal_row, column=6).value = new_grade
            ws_goal_table.cell(row=goal_row, column=7).value = str(out_cell.value)
            ws_goal_table.cell(row=goal_row, column=3).value = r

            local_index = ((r - 1) % SERIES_BLOCK_SIZE) + 1
            ws_goal_table.cell(row=goal_row, column=4).value = local_index
            dept = local_to_dept.get(local_index, "Unknown")
            ws_goal_table.cell(row=goal_row, column=5).value = dept

            logging.info(
                "Goal #%d improved by row %d (series %d): grade=%s, dept=%s",
                goal_idx + 1, r, series_num, new_grade, dept
            )

# ------------------ GPA ------------------

def grade_to_gpa(grade: str) -> float:
    return GRADE_TO_GPA.get(grade, 0.0)


def compute_weighted_gpa(ws_goal_table: Worksheet) -> float:
    weights = get_goal_weights(ws_goal_table)
    total = 0.0
    for i in range(NUM_GOALS):
        row = 2 + i
        best_grade = ws_goal_table.cell(row=row, column=2).value
        if best_grade in (None, ""):
            g = "F"
        else:
            g = str(best_grade).strip()
        score = grade_to_gpa(g)
        total += weights[i] * score
    return total


def all_goals_meet_min(ws_goal_table: Worksheet, min_grade: str) -> bool:
    if min_grade not in GRADE_INDEX:
        return True
    min_idx = GRADE_INDEX[min_grade]
    for i in range(NUM_GOALS):
        row = 2 + i
        best_grade = ws_goal_table.cell(row=row, column=2).value
        if best_grade in (None, ""):
            return False
        if GRADE_INDEX.get(str(best_grade).strip(), -1) < min_idx:
            return False
    return True

# ------------------ MAIN ------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python local_research.py local_research.xlsx [--oneseries|--continuous] [--debug] [--untilgpa X]")
        sys.exit(1)

    path = sys.argv[1]
    one_series_mode = ("--oneseries" in sys.argv)
    continuous_mode = ("--continuous" in sys.argv)
    debug_mode = ("--debug" in sys.argv)

    target_gpa: Optional[float] = None
    if "--untilgpa" in sys.argv:
        idx = sys.argv.index("--untilgpa")
        if idx + 1 >= len(sys.argv):
            print("ERROR: --untilgpa requires a numeric value, e.g. --untilgpa 3.5")
            sys.exit(1)
        try:
            target_gpa = float(sys.argv[idx + 1])
        except ValueError:
            print("ERROR: Invalid GPA value for --untilgpa.")
            sys.exit(1)

    if target_gpa is not None:
        one_series_mode = False
        continuous_mode = True

    if one_series_mode and continuous_mode:
        print("ERROR: Cannot use --oneseries and --continuous together.")
        sys.exit(1)

    setup_logging(debug_mode)
    logging.info("Starting with file: %s", path)

    if emergency_stop_check():
        logging.warning(".stop file detected before starting. Exiting.")
        print("Emergency stop file (.stop) found. Exiting without running.")
        sys.exit(0)

    # Load workbook
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        logging.error("Workbook not found: %s", path)
        print(f"ERROR: Workbook not found: {path}")
        sys.exit(1)
    except InvalidFileException:
        logging.error("Invalid or corrupted Excel file: %s", path)
        print(f"ERROR: Invalid or corrupted Excel file: {path}")
        sys.exit(1)
    except Exception as e:
        logging.error("Error loading workbook: %s", e)
        print(f"ERROR: Could not load workbook: {e}")
        sys.exit(1)

    # Sheets
    try:
        ws_goals = wb[SHEET_GOALS]
    except KeyError:
        logging.error("Sheet '%s' not found in workbook.", SHEET_GOALS)
        print(f"ERROR: Sheet '{SHEET_GOALS}' not found in workbook.")
        sys.exit(1)

    try:
        ws_series = wb[SHEET_SERIES]
    except KeyError:
        logging.error("Sheet '%s' not found in workbook.", SHEET_SERIES)
        print(f"ERROR: Sheet '{SHEET_SERIES}' not found in workbook.")
        sys.exit(1)

    ws_goal_table = ensure_goal_table_sheet(wb)
    ws_depts = ensure_dept_sheet(wb)

    generate_departments_and_tasks_if_needed(ws_goals, ws_depts)
    initialize_goals_if_needed(ws_goals, ws_goal_table)

    tasks = load_tasks_from_sheet4(ws_depts)
    if not tasks or len(tasks) < 19:
        logging.error("Could not load 19 tasks from Sheet4. Found %d tasks.", len(tasks))
        print("ERROR: Sheet4 does not contain 19 valid tasks. Please check or delete Sheet4 to regenerate.")
        sys.exit(1)

    departments, dept_last_index, local_to_dept = build_department_maps(tasks)

    try:
        while True:
            if emergency_stop_check():
                logging.warning(".stop file detected mid-run. Saving and exiting.")
                wb.save(path)
                print("Emergency stop requested. Progress saved. Exiting.")
                sys.exit(0)

            mission = get_latest_goal(ws_goals)
            if not mission:
                logging.error("No mission/goal found in Sheet1 column A.")
                print("ERROR: No mission/goal found in Sheet1 column A.")
                sys.exit(1)

            mode = detect_mission_mode(mission)
            print(f"Mission mode: {mode}")

            start_row, end_row, series_num = find_active_series_block(ws_series)
            logging.info("Processing series #%d (rows %d-%d).", series_num, start_row, end_row)
            print(f"Processing series #{series_num} (rows {start_row}-{end_row})...")

            ensure_column_a(ws_series, mission, tasks, departments, dept_last_index,
                            start_row, end_row, series_num)
            if emergency_stop_check():
                logging.warning(".stop detected after Column A. Saving and exiting.")
                wb.save(path)
                print("Emergency stop requested after Column A. Progress saved. Exiting.")
                sys.exit(0)

            ensure_column_b(ws_series, mission, tasks, departments, dept_last_index,
                            ws_goal_table,
                            start_row, end_row, series_num)
            if emergency_stop_check():
                logging.warning(".stop detected after Column B. Saving and exiting.")
                wb.save(path)
                print("Emergency stop requested after Column B. Progress saved. Exiting.")
                sys.exit(0)

            ensure_outputs_and_grades(ws_series, ws_goal_table, ws_depts,
                                      tasks, start_row, end_row, series_num, mode, mission)
            if emergency_stop_check():
                logging.warning(".stop detected after Column C/grades. Saving and exiting.")
                wb.save(path)
                print("Emergency stop requested after Column C/grades. Progress saved. Exiting.")
                sys.exit(0)

            # Check completion of the block
            block_complete = True
            for r in range(start_row, start_row + 19):
                if ws_series.cell(row=r, column=1).value in (None, ""):
                    block_complete = False
                    break
                if ws_series.cell(row=r, column=2).value in (None, ""):
                    block_complete = False
                    break
                if ws_series.cell(row=r, column=3).value in (None, ""):
                    block_complete = False
                    break
                for col in range(4, 4 + NUM_GOALS):
                    if ws_series.cell(row=r, column=col).value in (None, ""):
                        block_complete = False
                        break
                if not block_complete:
                    break
            if ws_series.cell(row=end_row, column=1).value in (None, ""):
                block_complete = False
            if ws_series.cell(row=end_row, column=2).value in (None, ""):
                block_complete = False

            if block_complete:
                logging.info("Series #%d fully filled. Updating best rows and backing up.", series_num)
                update_best_rows(ws_series, ws_goal_table, local_to_dept,
                                 start_row, end_row, series_num)
                wb.save(path)
                backup_workbook(path, series_num)
                print(f"Series #{series_num} completed. Backup created. Sheet3 updated (if improved rows found).")

                if target_gpa is not None:
                    gpa = compute_weighted_gpa(ws_goal_table)
                    logging.info("Current weighted GPA: %.3f (target %.3f)", gpa, target_gpa)
                    print(f"Current weighted GPA: {gpa:.3f} (target {target_gpa:.3f})")

                    if all_goals_meet_min(ws_goal_table, MIN_GOAL_GRADE) and gpa >= target_gpa:
                        logging.info("GPA target and per-goal minimum met. Stopping.")
                        print("GPA target and per-goal minimum met. Stopping.")
                        wb.save(path)
                        return

                if one_series_mode or not continuous_mode:
                    logging.info("Stopping after completing one series (oneseries or default mode).")
                    break
                else:
                    logging.info("Continuing to next series due to --continuous flag.")
                    continue
            else:
                wb.save(path)
                logging.info("Series #%d partially filled. Progress saved. Halting.", series_num)
                print(f"Series #{series_num} partially completed. Progress saved. Will resume next run.")
                break

        wb.save(path)
        logging.info("Run finished successfully.")
        print("Done.")

    except KeyboardInterrupt:
        logging.warning("KeyboardInterrupt received. Saving workbook and exiting.")
        wb.save(path)
        print("\nManual emergency stop (Ctrl+C). Progress saved. Exiting.")
        sys.exit(0)
    except Exception as e:
        logging.error("Unexpected error in main loop: %s", e)
        wb.save(path)
        print(f"ERROR: Unexpected error occurred: {e}. Progress saved.")
        sys.exit(1)


if __name__ == "__main__":
    main()'''

# ===== Engine T source =====
#===ENGINE_T_START===
T_CODE = r'''import os
import io
import json
import base64
import openpyxl
import matplotlib.pyplot as plt
from openai import OpenAI

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_MODEL = "gpt-4.1-mini"
MAX_ITERATIONS = 8

client = OpenAI(api_key=OPENAI_API_KEY)


def _safe_str(value):
    return "" if value is None else str(value).strip()


def read_prompt_from_sheet1(ws):
    """Read desired prompt from Sheet1, preferring A1 then A2 downward."""
    first = _safe_str(ws["A1"].value)
    if first and first.upper() not in {"W", "O", "R", "T", "H", "Y"}:
        return first

    lines = []
    for row in range(2, ws.max_row + 1):
        v = _safe_str(ws.cell(row=row, column=1).value)
        if not v:
            if lines:
                break
            continue
        lines.append(v)
    return "\n".join(lines).strip()


def get_or_create_next_sheet(wb):
    """Use the following page after Sheet1. Fallback to Sheet2."""
    if "Sheet1" in wb.sheetnames:
        idx = wb.sheetnames.index("Sheet1")
        if idx + 1 < len(wb.sheetnames):
            return wb[wb.sheetnames[idx + 1]]
    return wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.create_sheet("Sheet2")


def ensure_output_header(ws):
    if ws.max_row == 1 and ws["A1"].value is None:
        ws["A1"] = "Iteration"
        ws["B1"] = "Status"
        ws["C1"] = "EngineerPlanJSON"
        ws["D1"] = "Summary"


def load_previous_iterations(ws):
    records = []
    for row in range(2, ws.max_row + 1):
        iteration = ws.cell(row=row, column=1).value
        status = _safe_str(ws.cell(row=row, column=2).value)
        payload = _safe_str(ws.cell(row=row, column=3).value)
        summary = _safe_str(ws.cell(row=row, column=4).value)
        if iteration is None and not status and not payload and not summary:
            continue
        records.append({
            "iteration": iteration,
            "status": status,
            "payload": payload,
            "summary": summary,
        })
    return records


def ask_technical_engineer(prompt_text, prior_records, iteration):
    history = []
    for rec in prior_records[-6:]:
        history.append({
            "iteration": rec.get("iteration"),
            "status": rec.get("status"),
            "summary": rec.get("summary"),
        })

    prompt = f"""
You are T-Technical Engineer.

MISSION:
Engineer a solution to the user's prompt by iterating and testing ideas.
Focus on engineering artifacts: image visualization plans, graphs, simulations,
tradeoff analysis, technical architecture, and implementation experiments.

USER PROMPT:
{prompt_text}

CURRENT ITERATION:
{iteration}

PRIOR ITERATION LOG (vertical history from worksheet):
{json.dumps(history, indent=2)}

Return STRICT JSON with these keys:
- status: one of ["ITERATE", "SOLVED", "BLOCKED"]
- summary: short technical update
- bottlenecks: list of strings
- development_time: string estimate
- cost_estimate: string estimate
- risk_level: one of ["LOW", "MEDIUM", "HIGH"]
- feasibility: one of ["LOW", "MEDIUM", "HIGH"]
- technical_challenges: list of strings
- simulation_plan: concise simulation/test design
- graph_plan: concise graph/visualization plan
- next_actions: list of concrete engineering tasks

Important:
- If prior results are insufficient, set status to ITERATE.
- If solution is feasible and validated enough, set status to SOLVED.
- If impossible with current constraints, set status to BLOCKED.
JSON only.
"""

    response = client.chat.completions.create(
        model=OPENAI_MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )
    return response.choices[0].message.content


def parse_json_response(text):
    cleaned = (text or "").strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.strip("`")
        if cleaned.startswith("json"):
            cleaned = cleaned[4:].strip()
    return json.loads(cleaned)


def generate_iteration_plot(ws, upto_row):
    scores = []
    for row in range(2, upto_row + 1):
        risk = _safe_str(ws.cell(row=row, column=3).value)
        if not risk:
            continue
        try:
            payload = json.loads(risk)
        except Exception:
            continue
        fea = payload.get("feasibility", "MEDIUM")
        lvl = {"LOW": 1, "MEDIUM": 2, "HIGH": 3}.get(str(fea).upper(), 2)
        scores.append((row - 1, lvl))

    if not scores:
        return ""

    xs = [x for x, _ in scores]
    ys = [y for _, y in scores]

    fig, ax = plt.subplots(figsize=(6, 3))
    ax.plot(xs, ys, marker="o")
    ax.set_title("T-Technical Engineer Feasibility Trend")
    ax.set_xlabel("Iteration")
    ax.set_ylabel("Feasibility (1=Low,3=High)")
    ax.set_ylim(0.8, 3.2)
    ax.grid(True, alpha=0.3)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


def append_iteration(ws, iteration, payload):
    row = ws.max_row + 1
    ws.cell(row=row, column=1).value = iteration
    ws.cell(row=row, column=2).value = payload.get("status", "ITERATE")
    ws.cell(row=row, column=3).value = json.dumps(payload, ensure_ascii=False)
    ws.cell(row=row, column=4).value = payload.get("summary", "")
    return row


def main():
    file = "Worthy.xlsx"
    wb = openpyxl.load_workbook(file)

    if "Sheet1" not in wb.sheetnames:
        raise SystemExit("Sheet1 not found. Expected user prompt in Sheet1.")

    ws_prompt = wb["Sheet1"]
    ws_out = get_or_create_next_sheet(wb)
    ensure_output_header(ws_out)

    prompt_text = read_prompt_from_sheet1(ws_prompt)
    if not prompt_text:
        raise SystemExit("No prompt found in Sheet1 column A.")

    previous = load_previous_iterations(ws_out)
    solved_or_blocked = any(r.get("status") in {"SOLVED", "BLOCKED"} for r in previous)

    if solved_or_blocked:
        wb.save(file)
        print("T-Technical Engineer: prior run already reached SOLVED/BLOCKED.")
        return

    iteration = len(previous) + 1
    while iteration <= MAX_ITERATIONS:
        print(f"T-Technical Engineer iteration {iteration}...")

        raw = ask_technical_engineer(prompt_text, previous, iteration)
        payload = parse_json_response(raw)

        row = append_iteration(ws_out, iteration, payload)
        chart_b64 = generate_iteration_plot(ws_out, row)
        if chart_b64:
            ws_out.cell(row=row, column=5).value = chart_b64

        wb.save(file)

        status = str(payload.get("status", "ITERATE")).upper()
        previous = load_previous_iterations(ws_out)

        if status in {"SOLVED", "BLOCKED"}:
            print(f"T-Technical Engineer finished with status={status}.")
            break

        iteration += 1

    if iteration > MAX_ITERATIONS:
        print("T-Technical Engineer reached max iterations; last status kept as ITERATE.")


if __name__ == "__main__":
    main()'''

# ===== Engine H source =====
#===ENGINE_H_START===
H_CODE = r'''#!/usr/bin/env python3
import sys
import os
import shutil
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

import json
from openai import OpenAI

# Pull background passed in from the main Worthy controller (Sheet1!A3+)
try:
    BACKGROUND = WORTHY_BACKGROUND
except NameError:
    BACKGROUND = ""

client = OpenAI()

SHEET_QUESTIONS = "Questions"
SHEET_ANSWERS = "AnswersHistory"
SHEET_PROFILE = "Profile"
SHEET_PLAN = "Plan"

LOG_FILE = None
AUTOSAVE_EVERY = 25


def setup_logging(debug: bool):
    level = logging.DEBUG if debug else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler()],
    )
    logging.info("Logging initialized. Debug=%s", debug)


def emergency_stop_check() -> bool:
    return os.path.exists(".stop")


def backup_workbook(path: str):
    base, ext = os.path.splitext(path)
    backup_path = f"{base}_backup_latest{ext}"
    try:
        shutil.copyfile(path, backup_path)
        logging.info("Backup created: %s", backup_path)
    except Exception as e:
        logging.error("Failed to create backup: %s", e)


def load_or_create_workbook(path: str) -> Workbook:
    p = Path(path)
    if p.exists():
        try:
            wb = load_workbook(path)
            logging.info("Loaded existing workbook: %s", path)
            return wb
        except InvalidFileException:
            logging.error("Invalid or corrupted Excel file: %s", path)
            print(f"ERROR: Invalid or corrupted Excel file: {path}")
            sys.exit(1)
        except Exception as e:
            logging.error("Error loading workbook: %s", e)
            print(f"ERROR: Could not load workbook: {e}")
            sys.exit(1)
    else:
        logging.info("Workbook not found, creating new file: %s", path)
        wb = Workbook()
        return wb


def ensure_questions_sheet(wb: Workbook) -> Worksheet:
    if SHEET_QUESTIONS in wb.sheetnames:
        ws = wb[SHEET_QUESTIONS]
    else:
        ws = wb.active
        ws.title = SHEET_QUESTIONS
        ws.append([
            "Q_ID",
            "Domain",
            "QuestionText",
            "AnswerType",
            "ContextTag",
            "DependsOn_Q",
            "DependsOn_Value",
            "ActiveFlag",
            "Answer"
        ])
    return ws


def seed_example_questions(ws: Worksheet):
    if ws.max_row > 1:
        logging.info("Questions already present; skipping seeding.")
        return

    logging.info("Seeding example questions into Questions sheet.")

    rows = [
        ["Q001", "general", "Where were you primarily raised?", "list:midwest,west,east,south,outside_us", "", "", "", 1, ""],
        ["Q002", "general", "How would you best describe your current social class?", "list:working,middle,upper,unsure", "", "", "", 1, ""],
        ["Q003", "general", "Approximate population of the place you grew up?", "list:<5000,5k-50k,50k-250k,>250k", "", "", "", 1, ""],
        ["Q100", "family", "In your town, does family reputation noticeably affect your opportunities?", "YNMN", "SMALL_TOWN", "Q003", "<5000", 0, ""],
        ["Q101", "career", "Have you ever considered leaving your town to pursue better opportunities elsewhere?", "YNMN", "SMALL_TOWN", "Q003", "<5000", 0, ""],
        ["Q200", "career", "Are you willing to relocate for a significantly better job opportunity?", "YNMN", "", "", "", 1, ""],
        ["Q201", "career", "Do you prefer stable long-term employment over high-risk high-reward paths?", "YNMN", "", "", "", 1, ""],
        ["Q202", "career", "How satisfied are you with your current level of skill growth?", "list:very_satisfied,satisfied,neutral,unsatisfied,very_unsatisfied", "", "", "", 1, ""],
        ["Q300", "mental", "Do you frequently ruminate about past decisions?", "YNMN", "", "", "", 1, ""],
        ["Q301", "mental", "Do you have at least one person you can be fully honest with?", "YNMN", "", "", "", 1, ""],
        ["Q302", "mental", "On most weeks, do you have at least one block of time that feels genuinely restorative?", "YNMN", "", "", "", 1, ""],
        ["Q400", "financial", "Do you currently use a written or digital monthly budget that you follow?", "YNMN", "", "", "", 1, ""],
        ["Q401", "financial", "Do you carry any high-interest debt (e.g., credit cards)?", "YNMN", "", "", "", 1, ""],
        ["Q402", "financial", "Roughly how many months of essential expenses could you cover with your savings?", "list:<1,1-3,3-6,>6,unsure", "", "", "", 1, ""],
    ]

    for r in rows:
        ws.append(r)


def apply_dropdowns(ws: Worksheet):
    logging.info("Applying data validation dropdowns to Questions sheet.")
    max_row = ws.max_row

    for row in range(2, max_row + 1):
        answer_type = ws.cell(row=row, column=4).value
        if not answer_type:
            continue

        dv: Optional[DataValidation] = None

        if answer_type == "YNMN":
            dv = DataValidation(
                type="list",
                formula1='"Yes,No,Maybe,Not sure"',
                allow_blank=True
            )
        elif answer_type.startswith("list:"):
            options = answer_type.split(":", 1)[1]
            dv = DataValidation(
                type="list",
                formula1=f'"{options}"',
                allow_blank=True
            )

        if dv is not None:
            ws.add_data_validation(dv)
            cell = ws.cell(row=row, column=9)
            dv.add(cell)


def ensure_answers_sheet(wb: Workbook) -> Worksheet:
    if SHEET_ANSWERS in wb.sheetnames:
        ws = wb[SHEET_ANSWERS]
    else:
        ws = wb.create_sheet(SHEET_ANSWERS)
        ws.append(["SessionID", "Timestamp", "Q_ID", "Answer"])
    return ws


def ensure_profile_sheet(wb: Workbook) -> Worksheet:
    if SHEET_PROFILE in wb.sheetnames:
        ws = wb[SHEET_PROFILE]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(SHEET_PROFILE)
    ws.append(["Trait", "Value", "Notes"])
    return ws


def ensure_plan_sheet(wb: Workbook) -> Worksheet:
    if SHEET_PLAN in wb.sheetnames:
        ws = wb[SHEET_PLAN]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(SHEET_PLAN)
    ws.append(["Domain", "StepOrder", "TimeHorizon", "Action", "Rationale", "GameTheoryView"])
    return ws


def get_answers_dict(ws_q: Worksheet) -> Dict[str, Any]:
    answers: Dict[str, Any] = {}
    for row in range(2, ws_q.max_row + 1):
        qid = ws_q.cell(row=row, column=1).value
        ans = ws_q.cell(row=row, column=9).value
        if qid:
            answers[str(qid)] = ans
    return answers


def compute_context_flags(ws_q: Worksheet):
    logging.info("Computing context flags (ActiveFlag) based on current answers.")
    answers = get_answers_dict(ws_q)

    small_town = (answers.get("Q003") == "<5000")

    for row in range(2, ws_q.max_row + 1):
        context_tag = ws_q.cell(row=row, column=5).value
        depends_on_q = ws_q.cell(row=row, column=6).value
        depends_on_val = ws_q.cell(row=row, column=7).value
        active_cell = ws_q.cell(row=row, column=8)

        if not context_tag and not depends_on_q:
            if active_cell.value is None:
                active_cell.value = 1
            continue

        if context_tag == "SMALL_TOWN" and depends_on_q == "Q003" and depends_on_val == "<5000":
            active_cell.value = 1 if small_town else 0
        else:
            if active_cell.value is None:
                active_cell.value = 0


import uuid
import datetime as dt


def log_answers(ws_q: Worksheet, ws_a: Worksheet) -> str:
    session_id = str(uuid.uuid4())
    now = dt.datetime.now().isoformat(timespec="seconds")
    logging.info("Logging answers for new session: %s", session_id)

    count = 0
    for row in range(2, ws_q.max_row + 1):
        active = ws_q.cell(row=row, column=8).value
        qid = ws_q.cell(row=row, column=1).value
        ans = ws_q.cell(row=row, column=9).value
        if not qid or not active or ans in (None, ""):
            continue

        ws_a.append([session_id, now, qid, ans])
        count += 1

        if count % AUTOSAVE_EVERY == 0:
            logging.info("Logged %d answers so far in this session.", count)

    logging.info("Total answers logged this session: %d", count)
    return session_id


def infer_dimensions_from_background(background: str,
                                     answers: Dict[str, Any]) -> List[Dict[str, str]]:
    if not background or not str(background).strip():
        logging.info("No BACKGROUND text available; skipping background-based inference.")
        return []

    compact_answers = {k: v for k, v in answers.items() if v not in (None, "")}

    prompt = f"""
You are building a concise, non-pathologizing life profile for one person.

You will infer only **high-level tendencies** and contextual influences.
Do NOT invent or state clinical diagnoses (e.g., "depression", "ADHD", "bipolar").
Instead, talk about *influences*, *stressors*, *protective factors*, and *tendencies*.

BACKGROUND NOTES (free-form, messy, from an Excel sheet):
"""{background}"""

KNOWN ANSWERS FROM A SEPARATE SURVEY (if any):
{json.dumps(compact_answers, indent=2)}

Using only what is reasonably implied:

1. Focus especially on:
   - PsychometricSignals
   - SociologicalFactors
   - PersonalityTraits
   - CulturalBackground
   - FamilyExpectations
   - SocioeconomicGoals
   - MentalHealthInfluences
   - WorkStyle
   - LearningStyle
   - Values
   - RelationshipPatterns
   - StrategicPosture
   - EquilibriumForecast

2. For each dimension where you have enough signal, create **2–4 short bullets**.
3. If there really is not enough information for a dimension, skip it.

Return a **JSON list** of objects. Each object must have:

- "dimension": one of
  ["PsychometricSignals","SociologicalFactors","PersonalityTraits","CulturalBackground","FamilyExpectations",
   "SocioeconomicGoals","MentalHealthInfluences","WorkStyle",
   "LearningStyle","Values","RelationshipPatterns","StrategicPosture","EquilibriumForecast"]
- "label": a short 2–6 word label
- "explanation": 1–2 sentences explaining what you are inferring and why.

Answer with JSON only, no commentary.
"""

    try:
        logging.info("Calling OpenAI for background-based profile inference.")
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
        )
        raw = resp.choices[0].message.content
        raw = raw.strip()
        data = json.loads(raw)
        if not isinstance(data, list):
            logging.warning("Background inference: JSON root is not a list, ignoring.")
            return []
        result = []
        for item in data:
            dim = str(item.get("dimension", "")).strip()
            label = str(item.get("label", "")).strip()
            expl = str(item.get("explanation", "")).strip()
            if not dim or not label or not expl:
                continue
            result.append({
                "dimension": dim,
                "label": label,
                "explanation": expl,
            })
        logging.info("Background-based inference produced %d items.", len(result))
        return result
    except Exception as e:
        logging.error("Error during background-based inference: %s", e)
        return []


def infer_equilibrium_steps(background: str,
                            answers: Dict[str, Any],
                            profile_rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    compact_answers = {k: v for k, v in answers.items() if v not in (None, "")}
    profile_payload = [row for row in profile_rows if row.get("trait")]

    prompt = f"""
You are a predictive human-factors modeler.

Goal:
- Infer what this person is likely like (psychometric + sociological + strategic posture).
- Predict likely pushback and strategic responses from other actors.
- Use game-theory framing and forecast a likely equilibrium path.

Rules:
- Create a sequence of exactly 8 steps.
- Think one step at a time.
- At each step, read all prior steps and remain consistent with them.
- Each step must include:
  1) actor
  2) action
  3) rationale
  4) impact_on_equilibrium
- Keep each field concise and non-clinical.
- Do not diagnose mental illnesses.

BACKGROUND NOTES:
\"\"\"{background}\"\"\"

KNOWN ANSWERS:
{json.dumps(compact_answers, indent=2)}

INFERRED PROFILE:
{json.dumps(profile_payload, indent=2)}

Return JSON only in this exact schema:
[
  {{
    "step": 1,
    "actor": "...",
    "action": "...",
    "rationale": "...",
    "impact_on_equilibrium": "...",
    "domain": "General|Career|Financial|Mental|Relationships"
  }}
]
"""

    try:
        logging.info("Calling OpenAI for equilibrium-step inference.")
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
        )
        raw = (resp.choices[0].message.content or "").strip()
        data = json.loads(raw)
        if not isinstance(data, list):
            return []

        steps: List[Dict[str, str]] = []
        for item in data:
            step_num = item.get("step")
            actor = str(item.get("actor", "")).strip()
            action = str(item.get("action", "")).strip()
            rationale = str(item.get("rationale", "")).strip()
            impact = str(item.get("impact_on_equilibrium", "")).strip()
            domain = str(item.get("domain", "General")).strip() or "General"
            if not isinstance(step_num, int):
                continue
            if not actor or not action or not rationale or not impact:
                continue
            steps.append({
                "step": step_num,
                "actor": actor,
                "action": action,
                "rationale": rationale,
                "impact_on_equilibrium": impact,
                "domain": domain,
            })
        steps.sort(key=lambda x: x["step"])
        return steps[:8]
    except Exception as e:
        logging.error("Error during equilibrium-step inference: %s", e)
        return []


def build_profile(ws_q: Worksheet, ws_profile: Worksheet):
    logging.info("Rebuilding profile from answers + background.")

    if ws_profile.max_row > 1:
        ws_profile.delete_rows(2, ws_profile.max_row)

    ans = get_answers_dict(ws_q)

    try:
        inferred = infer_dimensions_from_background(BACKGROUND, ans)
    except Exception as e:
        logging.error("Failed background-based inference: %s", e)
        inferred = []

    for item in inferred:
        ws_profile.append([
            item["dimension"],
            item["label"],
            item["explanation"],
        ])

    stable = ans.get("Q201")
    if stable == "Yes":
        ws_profile.append([
            "RiskTolerance", "Low",
            "You report preferring stable, predictable paths over high-risk plays."
        ])
    elif stable == "No":
        ws_profile.append([
            "RiskTolerance", "High",
            "You report being open to higher risk in exchange for upside."
        ])
    else:
        ws_profile.append([
            "RiskTolerance", "Medium",
            "Your answers suggest a mixed or uncertain stance on risk vs. stability."
        ])

    ruminate = ans.get("Q300")
    if ruminate == "Yes":
        ws_profile.append([
            "Rumination", "High",
            "Frequent rumination can drain mental energy and slow decisions."
        ])
    elif ruminate == "No":
        ws_profile.append([
            "Rumination", "Lower",
            "You report less frequent rumination, which may free up mental bandwidth."
        ])
    else:
        ws_profile.append([
            "Rumination", "Unclear",
            "Your current answers do not give a clear picture of your rumination level."
        ])

    honest_person = ans.get("Q301")
    if honest_person == "Yes":
        ws_profile.append([
            "TrustedConnection", "Present",
            "You have at least one person with whom you can be fully honest."
        ])
    elif honest_person == "No":
        ws_profile.append([
            "TrustedConnection", "Missing",
            "You report lacking a fully honest connection; building one could be high leverage."
        ])
    else:
        ws_profile.append([
            "TrustedConnection", "Unclear",
            "Your support network picture is incomplete from current answers."
        ])

    logging.info("Profile rebuilt: %d rows total.", ws_profile.max_row - 1)


def build_plan(ws_profile: Worksheet, ws_plan: Worksheet):
    logging.info("Building plan from profile.")
    profile_rows: List[Dict[str, str]] = []

    for row in range(2, ws_profile.max_row + 1):
        trait = ws_profile.cell(row=row, column=1).value
        label = ws_profile.cell(row=row, column=2).value
        notes = ws_profile.cell(row=row, column=3).value

        if not trait:
            continue

        trait_str = str(trait)
        label_str = str(label or "")
        notes_str = str(notes or "")
        profile_rows.append({
            "trait": trait_str,
            "label": label_str,
            "notes": notes_str,
        })

        if trait_str in ["SocioeconomicGoals", "Financial", "RiskTolerance"]:
            domain = "Financial"
        elif trait_str in ["WorkStyle", "Career", "SkillGrowth"]:
            domain = "Career"
        elif trait_str in ["MentalHealthInfluences", "Rumination"]:
            domain = "Mental"
        elif trait_str in ["FamilyExpectations", "CulturalBackground", "RelationshipPatterns", "TrustedConnection"]:
            domain = "Relationships"
        else:
            domain = "General"

        action = f"Design one concrete experiment aligned with '{label_str}' in the {domain} domain."
        rationale = f"Based on profile dimension {trait_str}: {notes_str}"
        game_view = "Treat this as a small, low-stakes iteration rather than a permanent life commitment."

        ws_plan.append([
            domain,
            ws_plan.max_row,
            "3-6 months",
            action,
            rationale,
            game_view,
        ])

    equilibrium_steps = infer_equilibrium_steps(BACKGROUND, get_answers_dict(ws_profile.parent[SHEET_QUESTIONS]), profile_rows)
    for item in equilibrium_steps:
        ws_plan.append([
            item["domain"],
            item["step"],
            "near-term equilibrium path",
            f"{item['actor']}: {item['action']}",
            item["rationale"],
            item["impact_on_equilibrium"],
        ])

    logging.info("Plan building complete: %d rows.", ws_plan.max_row - 1)


def main():
    if len(sys.argv) < 2:
        print("Usage: python life_planner.py Worthy.xlsx [--debug]")
        sys.exit(1)

    path = sys.argv[1]
    debug = ("--debug" in sys.argv)

    setup_logging(debug)

    if emergency_stop_check():
        logging.warning(".stop file detected before starting. Exiting.")
        print("Emergency stop requested before start. Exiting.")
        sys.exit(0)

    wb = load_or_create_workbook(path)

    ws_q = ensure_questions_sheet(wb)
    seed_example_questions(ws_q)
    apply_dropdowns(ws_q)

    ws_a = ensure_answers_sheet(wb)
    ws_profile = ensure_profile_sheet(wb)
    ws_plan = ensure_plan_sheet(wb)

    compute_context_flags(ws_q)
    session_id = log_answers(ws_q, ws_a)

    build_profile(ws_q, ws_profile)
    build_plan(ws_profile, ws_plan)

    backup_workbook(path)
    wb.save(path)

    print(f"Life planner (Engine H) run complete. SessionID={session_id}")


if __name__ == "__main__":
    main()
'''

# ===== Engine Y source =====
#===ENGINE_Y_START===
Y_CODE = r'''import openpyxl
import re
from statistics import mean
from openai import OpenAI

client = OpenAI()

def _sheet_or_none(wb, name):
    return wb[name] if name in wb.sheetnames else None


def load_signal_sources(xlsx_path="Worthy.xlsx"):
    wb = openpyxl.load_workbook(xlsx_path)
    sheet1 = wb["Sheet1"]
    sheet2 = wb["Sheet2"]
    profile = _sheet_or_none(wb, "Profile")
    plan = _sheet_or_none(wb, "Plan")
    answers = _sheet_or_none(wb, "AnswersHistory")

    # Base source: single raw entry in Sheet1!A1 if user is directly providing text there.
    raw_a1 = (sheet1["A1"].value or "").strip()
    selector_tokens = {"W", "O", "R", "T", "H", "Y"}
    seed_text = raw_a1 if raw_a1.upper() not in selector_tokens else ""

    # Supplemental source: H outputs when available.
    h_signals = []
    if profile:
        for row in profile.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1] not in (None, ""):
                h_signals.append(f"{row[0]}: {row[1]}")
    if plan:
        for row in plan.iter_rows(min_row=2, values_only=True):
            if row and row[3]:
                rationale = row[4] if len(row) > 4 and row[4] else ""
                h_signals.append(f"PlanAction: {row[3]} | Rationale: {rationale}")
    if answers:
        for row in answers.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 4 and row[2] and row[3]:
                h_signals.append(f"Q: {row[2]} | A: {row[3]}")

    # Always-available second method: ingest existing Sheet1 information.
    sample_texts = []
    for row in sheet1.iter_rows(min_row=2, max_col=1, values_only=True):
        cell = row[0]
        if cell and str(cell).strip():
            sample_texts.append(str(cell).strip())

    if seed_text:
        sample_texts.insert(0, seed_text)
    if h_signals:
        sample_texts.extend(h_signals)

    return wb, sheet2, sample_texts, h_signals


def extract_traits(texts):
    texts = [t for t in texts if t and t.strip()]
    if not texts:
        texts = ["Direct, practical, reflective, and strategic communication style."]

    sentence_lengths, word_counts = [], []
    vocabulary = set()
    emotional_words = direct_flags = formal_flags = humor_flags = 0

    emotional_terms = {"feel", "emotion", "believe", "care", "love", "hate", "worried"}
    direct_terms = {"must", "should", "definitely", "clearly", "truth"}
    formal_terms = {"therefore", "consequently", "thus", "hence"}
    humor_terms = {"lol", "haha", "funny", "sarcasm"}

    for text in texts:
        words = [w for w in text.split() if w.strip()]
        if not words:
            continue
        sentences = re.split(r"[.!?]", text)
        sentence_lengths.extend([len(s.split()) for s in sentences if s.strip()])
        word_counts.append(len(words))
        vocabulary.update(w.lower() for w in words)

        for w in words:
            wl = re.sub(r"[^a-z]", "", w.lower())
            if wl in emotional_terms:
                emotional_words += 1
            if wl in direct_terms:
                direct_flags += 1
            if wl in formal_terms:
                formal_flags += 1
            if wl in humor_terms:
                humor_flags += 1

    total_words = max(sum(word_counts), 1)
    total_texts = max(len(texts), 1)
    avg_sentence = mean(sentence_lengths) if sentence_lengths else 12

    return {
        "VocabularyDensity": min(int((len(vocabulary) / total_words) * 1000), 100),
        "AverageSentenceLength": min(int(avg_sentence * 4), 100),
        "Directness": min(int((direct_flags / total_texts) * 200), 100),
        "Formality": min(int((formal_flags / total_texts) * 200), 100),
        "EmotionalTransparency": min(int((emotional_words / total_words) * 3000), 100),
        "HumorUsage": min(int((humor_flags / total_texts) * 200), 100),
        "StrategicThinking": 90,
        "ConsistencyOfVoice": 92,
    }


def build_engram_prompt(persona, h_signals, texts):
    traits = "\n".join([f"- {k}: {v}%" for k, v in persona.items()])
    h_block = "\n".join([f"- {x}" for x in h_signals[:40]]) if h_signals else "- None available"
    examples = "\n".join([f"- {t[:220]}" for t in texts[:20]])

    return f"""Create a high-fidelity personality engram for one target person.

SOURCE PRIORITY:
1) Use H-module signals as supplements when available.
2) Always use direct text evidence from workbook entries as the second method.

PERSONA VECTOR:
{traits}

H-MODULE SIGNALS:
{h_block}

EVIDENCE SNIPPETS:
{examples}

Design prompts that make replication robust across:
- Vocal cadence / rhythm
- Word choices and sentence structure
- Thought process and decision logic
- Regional dialect and idioms
- Values, beliefs, worldview anchors
- Known facts and biographical constraints

Output format:
Return a numbered list of prompts. Each prompt should be self-contained and deployable."""


def write_prompts_to_sheet2(sheet2, engram_text):
    # Reset only output area in A:C to keep behavior deterministic.
    for r in range(1, sheet2.max_row + 1):
        for c in range(1, 4):
            sheet2.cell(row=r, column=c).value = None

    sheet2["A1"] = "PromptID"
    sheet2["B1"] = "EngramPrompt"
    sheet2["C1"] = "Source"

    lines = [ln.strip() for ln in engram_text.splitlines() if ln.strip()]
    row = 2
    prompt_num = 1
    for ln in lines:
        normalized = re.sub(r"^\d+[\).\-\s]+", "", ln).strip()
        if len(normalized) < 10:
            continue
        sheet2.cell(row=row, column=1).value = f"Y-P{prompt_num:03d}"
        sheet2.cell(row=row, column=2).value = normalized
        sheet2.cell(row=row, column=3).value = "Y-Engram"
        row += 1
        prompt_num += 1

    return prompt_num - 1


if __name__ == "__main__":
    wb, sheet2, texts, h_signals = load_signal_sources("Worthy.xlsx")
    persona = extract_traits(texts)
    engram_request = build_engram_prompt(persona, h_signals, texts)

    completion = client.chat.completions.create(
        model="gpt-4.1",
        messages=[
            {"role": "system", "content": "You are an expert persona modeling and prompt-engineering architect."},
            {"role": "user", "content": engram_request},
        ],
    )
    engram_text = completion.choices[0].message.content or ""
    prompt_count = write_prompts_to_sheet2(sheet2, engram_text)
    wb.save("Worthy.xlsx")

    with open("YOU_PROMPT.txt", "w", encoding="utf-8") as f:
        f.write(engram_request + "\n\n---\n\n" + engram_text)

    print(f"Y engine complete. Wrote {prompt_count} engram prompts to Sheet2.")'''

# ---------------------------------------------------------------------------
# Engine runners
# ---------------------------------------------------------------------------

def _exec_engine(code_str: str, argv, background: str):
    """
    Execute an engine script in an isolated namespace.
    `argv` is a full list (e.g. ["worthy_R", "Worthy.xlsx", "--untilgpa", "3.4"])
    """
    import sys

    ns = {}
    ns["__name__"] = "__main__"
    ns["WORTHY_BACKGROUND"] = background

    # Send argv list directly to the engine
    sys.argv = argv

    exec(code_str, ns, ns)


def run_W(background: str):
    """Run Writer engine (W)."""
    _exec_engine(W_CODE, "worthy_W", background)


def run_O(background: str):
    """Run Operations / Control-Sheet engine (O)."""
    _exec_engine(O_CODE, ["worthy_O", WORTHY_XLSX_PATH], background)


def _exec_engine(code_str: str, argv_list, background: str):
    """
    Execute an engine script in isolated namespace.
    argv_list is full sys.argv for that engine.
    """
    import sys

    ns = {}
    ns["__name__"] = "__main__"
    ns["WORTHY_BACKGROUND"] = background
    sys.argv = argv_list

    exec(code_str, ns, ns)


def run_R(background: str):
    """
    Research engine wrapper with:
    - CLI --untilgpa support
    - Excel fallback (Sheet1!B1)
    - GPA + timing log written to a dedicated sheet 'R_Log'
    """
    import sys
    import time
    import openpyxl

    # ----------- LOAD WORKBOOK -----------
    wb = openpyxl.load_workbook(WORTHY_XLSX_PATH)

    # Create / get log sheet
    if "R_Log" not in wb.sheetnames:
        log_ws = wb.create_sheet("R_Log")
        log_ws.append(["Series", "Start Time", "End Time", "Elapsed (sec)", "Weighted GPA"])
        wb.save(WORTHY_XLSX_PATH)
    else:
        log_ws = wb["R_Log"]

    ws = wb["Sheet1"]

    # ----------- ARG HANDLING -----------
    # CLI override
    cli_gpa = None
    if "--untilgpa" in sys.argv:
        try:
            i = sys.argv.index("--untilgpa")
            cli_gpa = float(sys.argv[i+1])
        except:
            pass

    argv = ["worthy_R", WORTHY_XLSX_PATH]

    if cli_gpa is not None:
        argv += ["--untilgpa", str(cli_gpa)]
    else:
        # Excel fallback
        excel_gpa = ws["B1"].value
        if excel_gpa not in (None, ""):
            try:
                excel_val = float(excel_gpa)
                argv += ["--untilgpa", str(excel_val)]
            except:
                pass

    # ----------- GPA CALCULATOR /\ -----------
   

def run_T(background: str):
    """Run T-Technical Engineer iterative engineering engine (T)."""
    _exec_engine(T_CODE, "worthy_T", background)


def run_H(background: str):
    """Run Life Planner / Psychometric System engine (H)."""
    _exec_engine(H_CODE, "worthy_H", background)


def run_Y(background: str):
    """Run Persona / YOU-Mode engine (Y)."""
    _exec_engine(Y_CODE, "worthy_Y", background)


# ---------------------------------------------------------------------------
# Top-level controller
# ---------------------------------------------------------------------------


def main():
    selector, background = get_selector_and_background()
    if not selector:
        raise SystemExit("Sheet1!A1 is empty. Put one of W/O/R/T/H/Y there.")

    engines = {
        "W": run_W,
        "O": run_O,
        "R": run_R,
        "T": run_T,
        "H": run_H,
        "Y": run_Y,
    }

    runner = engines.get(selector)
    if runner is None:
        raise SystemExit(
            f"Unknown engine selector {selector!r} in Sheet1!A1. "
            f"Expected one of W/O/R/T/H/Y."
        )

    # Safe-bet mode: run exactly one engine, then exit.
    runner(background)


if __name__ == "__main__":
    main()
